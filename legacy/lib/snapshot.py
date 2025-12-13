import json
import os
import re
from datetime import datetime
from legacy.lib.utils import (
    load_devices,
    show_version,
    show_resources,
    show_interface,
    show_mac_address_table,
    show_ip_route,
    show_arp,
    show_logg,
    connect_to_device,
)
from legacy.customer_context import get_customer_name
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from rich.console import Console

console = Console()


# ============================================================
# Regex helpers 
# ============================================================


def extract_temperature(text: str):
    """Best-effort temperature extraction from various Cisco outputs.

    Returns:
        int     -> temperature in Celsius
        "OK"    -> if only status without explicit value is available
        None    -> if nothing can be parsed
    """
    # Helper untuk menyaring angka suhu tidak masuk akal
    def _san(v):
        try:
            x = int(v)
        except Exception:
            return None
        return x if 1 <= x <= 120 else None  # buang 0 dan nilai liar

    # --- 1) Khusus C9300: SYSTEM OUTLET/INLET/HOTSPOT (ambil OUTLET dulu) ---
    m = re.search(r"SYSTEM\s+OUTLET\s+\d+\s+\S+\s+(\d+)\s+Celsius", text, re.IGNORECASE)
    if m:
        v = _san(m.group(1))
        if v is not None:
            return v

    m = re.search(r"SYSTEM\s+HOTSPOT\s+\d+\s+\S+\s+(\d+)\s+Celsius", text, re.IGNORECASE)
    if m:
        v = _san(m.group(1))
        if v is not None:
            return v

    m = re.search(r"SYSTEM\s+INLET\s+\d+\s+\S+\s+(\d+)\s+Celsius", text, re.IGNORECASE)
    if m:
        v = _san(m.group(1))
        if v is not None:
            return v

    # --- 2) Nilai eksplisit umum (IOS-XE/C3850, dsb) ---
    for pat in [
        r"System Temperature Value:\s*(\d+)\s*Degree\s+Celsius",
        r"Inlet Temperature Value:\s*(\d+)\s*Degree\s+Celsius",
        r"System Temperature:\s*(\d+)\s*Celsius",
        r"Air\s+outlet\s+(\d+)C",                   # 4500/6500
        r"Chassis\s+Temperature\s*=\s*(\d+)",       # 4507
    ]:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            v = _san(m.group(1))
            if v is not None:
                return v

    # --- 3) Tabel sensor bergaya Nexus (CurTemp), prioritas BACK/Back ---
    m = re.search(r"\bBACK\s+\d+\s+\d+\s+(\d+)", text, re.IGNORECASE)
    if m:
        v = _san(m.group(1))
        if v is not None:
            return v

    # Variasi: "Back (D3)   70  46  28  Ok"
    m = re.search(r"\bBack(?:\s+\([^)]+\))?\s+\d+\s+\d+\s+(\d+)", text, re.IGNORECASE)
    if m:
        v = _san(m.group(1))
        if v is not None:
            return v

    # Jika ada OUTLET dalam tabel (ambil yang terbesar)
    arr = re.findall(r"\bOUTLET\s+\d+\s+\d+\s+(\d+)", text, re.IGNORECASE)
    arr = [_san(x) for x in arr]
    arr = [x for x in arr if x is not None]
    if arr:
        return max(arr)

    # --- 4) Nexus/N5K pola "Outlet ... 35 ok" ---
    m = re.search(r"Outlet.*?(\d+)\s+ok", text, re.IGNORECASE)
    if m:
        v = _san(m.group(1))
        if v is not None:
            return v

    # --- 5) SYSTEM TEMPERATURE STATUS (Exhaust/Exhaust Left) ---
    if "SYSTEM TEMPERATURE STATUS" in text:
        m = re.search(
            r"Exhaust(?:\s+Left)?\s+temperature:\s+(\d+)\s+Celsius",
            text,
            re.IGNORECASE,
        )
        if m:
            v = _san(m.group(1))
            if v is not None:
                return v

    # Fallback inlet ".... 31 Ok"
    m = re.search(r"Inlet.*?(\d+)\s+Ok", text, re.IGNORECASE)
    if m:
        v = _san(m.group(1))
        if v is not None:
            return v

    # --- 6) C8300 & generik: banyak baris "Temp: ... X Celsius" → rata-rata wajar ---
    arr = re.findall(r"\bTemp[:\s][^:\n]*?(\d+)\s+Celsius", text, re.IGNORECASE)
    arr = [_san(x) for x in arr]
    arr = [x for x in arr if x is not None]
    if arr:
        return sum(arr) // len(arr)

    # --- 7) Fallback 'TEMPERATURE is OK' jika hanya status tanpa angka ---
    if re.search(r"\bSYSTEM TEMPERATURE is OK\b|\bTEMPERATURE is OK\b", text, re.IGNORECASE):
        return "OK"

    # Khusus IOS ISR (mis. CISCO1941/K9): SYSTEM TEMPERATURE STATUS
    m = re.search(r"Exhaust\s*Fan\s*temperature:\s*(\d+)\s*Celsius", text, re.IGNORECASE)
    if m:
        try:
            v = int(m.group(1))
            if 1 <= v <= 120:
                return v
        except Exception:
            pass

    return None


def _reduce_statuses(statuses):

    if not statuses:
        return None

    norm = []
    for s in statuses:
        if not s:
            continue
        su = str(s).strip().upper()
        if su in ("OK", "GOOD", "NORMAL"):
            norm.append("OK")
        elif su in ("BAD", "FAIL", "FAILED", "FAULT", "FAULTY", "WARNING"):
            norm.append("Bad")
        elif su in ("NOT PRESENT", "NOT_PRESENT", "NOTPRESENT"):
            norm.append("NOT PRESENT")
        else:
            norm.append(str(s).strip())

    if "Bad" in norm:
        return ["Bad"]

    has_ok = "OK" in norm
    has_np = "NOT PRESENT" in norm
    if has_ok and has_np:
        return ["OK", "NOT PRESENT"]
    if has_ok:
        return ["OK"]
    if has_np:
        return ["NOT PRESENT"]

    out = []
    for v in norm:
        if v not in out:
            out.append(v)
    return out


def extract_fan_status(text: str):
    statuses = []

    # --- Tambahan khusus N5K: blok tabel "Fan:" ---
    n5k_rows = re.findall(
        r"^\s*(?:Chassis-\d+|PS-\d+)\s+\S+\s+\S+\s+(ok|fail)\s*$",
        text,
        re.MULTILINE | re.IGNORECASE,
    )
    for tok in n5k_rows:
        statuses.append(tok)

    # Pola-pola umum
    if re.search(r"\bFAN\s+is\s+OK\b", text, re.IGNORECASE):
        statuses.append("OK")

    m = re.search(r"Fantray\s*:\s*(Good|Bad)", text, re.IGNORECASE)
    if m:
        statuses.append(m.group(1))

    for tok in re.findall(
        r"FAN\s+PS-\d+\s+is\s+(OK|FAIL|NOT PRESENT)", text, re.IGNORECASE
    ):
        statuses.append(tok)

    for tok in re.findall(
        r"^\s*\d+\s+\d+\s+\d+\s+(OK|FAIL)", text, re.MULTILINE | re.IGNORECASE
    ):
        statuses.append(tok)

    for tok in re.findall(
        r"PS\d\s+FAN\s+\d+\s+(GOOD|BAD|NOT PRESENT)", text, re.IGNORECASE
    ):
        statuses.append(tok)

    for tok in re.findall(
        r"Switch\s+\d+\s+FAN\s+\d+\s+is\s+(OK|NOT PRESENT)", text, re.IGNORECASE
    ):
        statuses.append(tok)

    # Nexus bebas: "Fan1(sys_fan1) ... Ok"
    for tok in re.findall(r"Fan\d.*?\s+(Ok|Fail)\b", text, re.IGNORECASE):
        statuses.append(tok)

    for tok in re.findall(r"fan-fail:\s+(OK|FAIL)", text, re.IGNORECASE):
        statuses.append(tok)

    for tok in re.findall(
        r"RPM:\s+fan\d+\s+\S+\s+(Normal|Warning|Fault)", text, re.IGNORECASE
    ):
        statuses.append(tok)

    m = re.search(r"System Fan\s+(OK|Faulty|Fail)", text, re.IGNORECASE)
    if m:
        statuses.append(m.group(1))

    for tok in re.findall(
        r"Fan\s+\d+\s+(OK|Fail|Faulty)", text, re.IGNORECASE
    ):
        statuses.append(tok)

    if re.findall(r"\bFan\s+\d+\s+OK\b", text, re.IGNORECASE):
        statuses.append("OK")

    if re.findall(r"\bFC FAN\d.*?Celsius\b", text, re.IGNORECASE):
        statuses.append("OK")

    return _reduce_statuses(statuses)


def extract_psu_status(text: str):
    statuses = []

    # Tabel SW PID ... Good/OK/Bad/Warning
    if "SW  PID" in text:
        for line in text.splitlines():
            if re.search(r"PWR|PSU|Built-in", line, re.IGNORECASE):
                if re.search(r"Good|OK", line, re.IGNORECASE):
                    statuses.append("OK")
                elif re.search(r"Bad|WARNING", line, re.IGNORECASE):
                    statuses.append("Bad")

    # C8300 / C8200: "P: In pwr ... Normal/Warning/Fault"
    for tok in re.findall(
        r"P:\s+(?:In\s+)?pwr\s+\S+\s+(Normal|Warning|Fault)", text, re.IGNORECASE
    ):
        statuses.append(tok)

    # C2960: POWER, RPS
    if re.search(r"POWER\s+is\s+OK", text, re.IGNORECASE):
        statuses.append("OK")
    if re.search(r"RPS\s+is\s+NOT\s+PRESENT", text, re.IGNORECASE):
        statuses.append("NOT PRESENT")

    # WS-C2950G-48-EI (bila fungsi helper tambahan tersedia)
    try:
        if re.search(r"\bWS-C2950G-48-EI\b", text, re.IGNORECASE) or re.search(
            r"Internal\s+POWER\s+supply\s+is", text, re.IGNORECASE
        ):
            # Fungsi optional; aman bila tidak ada
            parsed = extract_psu_status_2950(text)  # type: ignore[name-defined]
            if parsed:
                statuses.extend(parsed)
    except NameError:
        pass

    # C3750: Sys Pwr Good
    if re.search(r"SW\s+PID.*Sys Pwr", text, re.IGNORECASE) and "Good" in text:
        statuses.append("OK")

    # 4507R-E power table good/bad
    if "Chassis Type : WS-C4507R-E" in text and "Power" in text and "Model No" in text:
        for tok in re.findall(
            r"^PS\d+\s+\S+\s+.+?\s+(good|bad)",
            text,
            re.MULTILINE | re.IGNORECASE,
        ):
            statuses.append(tok)

    # 6509: "power-supply X power-output-fail: OK/FAIL"
    for tok in re.findall(
        r"power-supply \d+ power-output-fail:\s+(\S+)", text, re.IGNORECASE
    ):
        statuses.append(tok)

    # Nexus 5K: tabel "Power Supply: ... ok"
    for tok in re.findall(
        r"^\d+\s+\S+\s+\S+\s+\S+\s+\S+\s+(ok)",
        text,
        re.MULTILINE | re.IGNORECASE,
    ):
        statuses.append(tok)

    # "SYSTEM POWER SUPPLY STATUS"
    m = re.search(
        r"Power Supply .* Status:\s+(Normal|Fault|Bad)", text, re.IGNORECASE
    )
    if m:
        statuses.append(m.group(1))

    # Tabel PS\d ... good/bad
    for tok in re.findall(
        r"^PS\d+\s+\S+\s+\S+\s+\S+\s+(good|bad)",
        text,
        re.MULTILINE | re.IGNORECASE,
    ):
        statuses.append(tok)

    # Vin/Vout Normal → OK
    if re.findall(r"(?:Vin|Vout).*?Normal", text, re.IGNORECASE):
        statuses.append("OK")

    # Main Power Supply is AC → OK
    if re.search(r"Main Power Supply is AC", text, re.IGNORECASE):
        statuses.append("OK")

    return _reduce_statuses(statuses)


def extract_crc_interfaces(text: str):
   # Cari interface yang punya CRC error dari output 'show interfaces'.
   # Hanya interface dengan CRC > 0 

    results = []

    blocks = re.split(
        r"\n(?=[A-Za-z][A-Za-z0-9\/\.]*\s+is\s+.+line protocol is\s+)",
        text,
    )

    for block in blocks:
        lines = block.strip().splitlines()
        if not lines:
            continue

        header = lines[0]
        m = re.match(
            r"^(?P<intf>\S+)\s+is\s+(?P<link_status>[^,]+),\s+line protocol is\s+(?P<protocol_status>[^\n]+)",
            header,
            re.IGNORECASE,
        )
        if not m:
            continue

        intf_name = m.group("intf").strip()
        link_status = m.group("link_status").strip()
        protocol_status = m.group("protocol_status").strip()

        crc_match = re.search(r"(\d+)\s+CRC", block, re.IGNORECASE)
        if not crc_match:
            continue

        crc_val = crc_match.group(1).strip()
        if crc_val in ("", "0", "0.0"):
            continue

        desc_match = re.search(
            r"^\s*Description[: ]+(.+)$", block, re.IGNORECASE | re.MULTILINE
        )
        description = desc_match.group(1).strip() if desc_match else ""

        results.append(
            {
                "interface": intf_name,
                "crc": crc_val,
                "link_status": link_status,
                "protocol_status": protocol_status,
                "description": description,
            }
        )

    return results


def extract_logs(text: str):
    #Ambil baris log penting dari output show logging.
    #Heuristik: baris yang mengandung pola %FAC-<level>-MNEMONIC:
    
    logs = []
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        if re.search(r"%[A-Z0-9_-]-\d-[A-Z0-9_]+:", line):
            logs.append(line)
    return logs


# ============================================================
# Helper untuk parsing syslog → Timestamp & Severity
# ============================================================
def _parse_syslog_line(line: str):
    """Parse Cisco-style syslog line.

    Returns (timestamp, severity_text, message_only)
    """
    # Pola dengan timestamp di depan
    m = re.search(
        r"^(?P<ts>[A-Z][a-z]{2}\s+\d{1,2}\s+\d{2}:\d{2}:\d{2}(?:\.\d+)?)\s*:\s*%(?P<fac>[A-Z0-9_-]+)-(?P<lvl>\d)-(?P<mnem>[A-Z0-9_]+):\s*(?P<msg>.*)$",
        line,
    )
    ts = None
    lvl = None
    msg = None

    if m:
        ts = m.group("ts").strip()
        lvl = m.group("lvl")
        msg = m.group("msg").strip()
    else:
        # Fallback: tanpa timestamp di depan
        m2 = re.search(
            r"%(?P<fac>[A-Z0-9_-]+)-(?P<lvl>\d)-(?P<mnem>[A-Z0-9_]+):\s*(?P<msg>.*)",
            line,
        )
        if not m2:
            return None, None, line
        lvl = m2.group("lvl")
        msg = m2.group("msg").strip()

    level_map = {
        "0": "EMERGENCY",
        "1": "ALERT",
        "2": "CRITICAL",
        "3": "ERROR",
        "4": "WARNING",
        "5": "NOTICE",
        "6": "INFO",
        "7": "DEBUG",
    }

    severity = level_map.get(str(lvl), None)
    return ts, severity, msg or line


# ============================================================
# Core snapshot logic
# ============================================================


def capture_device_output(creds, progress_callback=None):
    hostname = creds["hostname"]
    device_type = creds["os"]
    conn = connect_to_device(creds)

    if conn:
        msg = f"Connected to {hostname} ({device_type})..."
        if progress_callback:
            try:
                progress_callback(msg)
            except Exception:
                pass
        else:
            console.print(f"[bold cyan]{msg}[/bold cyan]")

        # ------------------------------------------------------------
        # Collect structured data via helper functions
        # ------------------------------------------------------------
        show_ver = show_version(conn, device_type)
        resources = show_resources(conn, device_type)
        interfaces = show_interface(conn)
        mac_address = show_mac_address_table(conn)
        ip_routes = show_ip_route(conn, device_type)
        arp_table = show_arp(conn, device_type)
        loggs = show_logg(conn, device_type)

        # ------------------------------------------------------------
        # Extra raw outputs for regex-based parsing (TEMP / FAN / PSU / LOGS)
        # ------------------------------------------------------------
        env_chunks = []
        for cmd in [
            "show environment all",
            "show environment",
            "show env all",
            "show env",
            "show environment power",
            "show environment fan",
        ]:
            try:
                out = conn.send_command(cmd)
                # Skip jika perangkat tidak support (Invalid input)
                if out and "Invalid input" not in out:
                    env_chunks.append(out)
            except Exception:
                continue

        env_text = "\n".join(env_chunks)

        # show logging penuh untuk kebutuhan sheet Logs
        try:
            raw_logging = conn.send_command("show logging")
        except Exception:
            raw_logging = ""

        # Parse dengan regex helper
        temperature = extract_temperature(env_text) if env_text else None
        fan_status = extract_fan_status(env_text) if env_text else None
        psu_status = extract_psu_status(env_text) if env_text else None

        parsed_logs = extract_logs(raw_logging) if raw_logging else [] # type: ignore
        if not parsed_logs and loggs:
            # fallback: pakai subset dari show_logg
            parsed_logs = extract_logs("\n".join(loggs))

        data = {
            "health_check": {
                "hostname": show_ver.get("hostname", hostname),
                "uptime": show_ver.get("uptime", ""),
                "version": show_ver.get("version", ""),
                "cpu_utilization": resources.get("cpu_utilization", ""),
                "memory_utilization": resources.get("memory_utilization", ""),
                "storage_utilization": resources.get("storage_utilization", ""),
                # nilai tambahan untuk sheet Health Check
                "temperature": temperature,
                "fan_status": fan_status,
                "psu_status": psu_status,
            },
            "interfaces": interfaces,
            "mac_address_table": mac_address,
            "routing_table": ip_routes,
            "arp_table": arp_table,
            "logs": loggs,
            # simpan hasil parsing regex log untuk sheet Logs
            "parsed_logs": parsed_logs,
        }

        return data

    else:
        msg = f"ERROR: Failed to capture from {hostname}"
        if progress_callback:
            try:
                progress_callback(msg)
            except Exception:
                pass
        else:
            console.print(f"[red]{msg}[/red]")


def autosize_columns(ws: Worksheet) -> None:
    """Autosize all columns in a worksheet based on content length."""
    for col in ws.columns:
        first_cell = col[0]
        if first_cell.column is None:
            continue

        col_letter = get_column_letter(first_cell.column)
        max_len = 0

        for cell in col:
            try:
                val = "" if cell.value is None else str(cell.value)
                if len(val) > max_len:
                    max_len = len(val)
            except Exception:
                pass

        ws.column_dimensions[col_letter].width = max_len + 2


def _status_to_text(val):
    """Convert fan/psu status (list/str/None) menjadi string rapi."""
    if isinstance(val, list):
        return ", ".join(str(v) for v in val if v is not None)
    if val is None:
        return ""
    return str(val)

# Summary CRC untuk sheet Health Check.
def _compute_crc_summary(interfaces):
   
    if not interfaces:
        return ""

    has_crc_error = False

    for intf in interfaces:
        crc_raw = intf.get("crc")
        if crc_raw in (None, ""):
            continue

        try:
            crc_int = int(str(crc_raw).strip())
        except ValueError:
            # kalau format CRC aneh, skip saja
            continue

        if crc_int > 0:
            has_crc_error = True
            break

    return "NOK" if has_crc_error else "OK"



def health_check(customer_name, data, base_dir, progress_callback=None):
    path = os.path.join(base_dir, "health_check")
    os.makedirs(path, exist_ok=True)

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    health_check_path = os.path.join(
        path, f"{customer_name}_health_check_{timestamp}.xlsx"
    )

    wb: Workbook = Workbook()

    # ---------------------------
    # Sheet 1: Health Check
    # ---------------------------
    ws_health: Worksheet = wb.create_sheet("Health Check", 0)

    headers_health = [
        "Hostname",
        "Version",
        "Cpu utilization",
        "Memory utilization",
        "Storage utilization",
        "Uptime",
        "PSU",
        "FAN",
        "Temperature (C)",
        "CRC",
    ]
    ws_health.append(headers_health)

    for hostname, device_data in data.items():
        health = device_data.get("health_check", {}) or {}
        interfaces = device_data.get("interfaces", []) or []

        # --- Hitung CRC summary langsung di sini ---
        crc_status = ""
        if interfaces:
            has_crc_error = False
            for intf in interfaces:
                crc_raw = intf.get("crc")
                if crc_raw in (None, ""):
                    continue

                try:
                    crc_int = int(str(crc_raw).strip())
                except ValueError:
                    # Kalau format CRC aneh, skip saja
                    continue

                if crc_int > 0:
                    has_crc_error = True
                    break

            crc_status = "NOK" if has_crc_error else "OK"

        # --- Temperature cell formatting ---
        temp_val = health.get("temperature", "")
        if isinstance(temp_val, list):
            temp_cell = ", ".join(str(v) for v in temp_val if v is not None)
        else:
            temp_cell = temp_val if temp_val is not None else ""

        row = [
            health.get("hostname", hostname),
            health.get("version", ""),
            health.get("cpu_utilization", ""),
            health.get("memory_utilization", ""),
            health.get("storage_utilization", ""),
            health.get("uptime", ""),
            _status_to_text(health.get("psu_status") or health.get("psu")),
            _status_to_text(health.get("fan_status") or health.get("fan")),
            temp_cell,
            crc_status,  # <- sekarang jelas OK / NOK
        ]
        ws_health.append(row)


    autosize_columns(ws_health)

    # ---------------------------
    # Sheet 2: CRC Interfaces (sudah ada, tetap dipakai)
    # ---------------------------
    ws_crc: Worksheet = wb.create_sheet("CRC Interfaces", 1)

    headers_crc = [
        "Hostname",
        "Interface",
        "CRC",
        "Link status",
        "Protocol status",
        "Description",
    ]
    ws_crc.append(headers_crc)

    current_row = 2

    for hostname, device_data in data.items():
        interfaces = device_data.get("interfaces", []) or []

        first_row_for_host = None
        rows_for_this_host = 0

        for intf in interfaces:
            crc_raw = intf.get("crc", "")
            crc = "" if crc_raw is None else str(crc_raw).strip()

            if crc in ("", "0"):
                continue

            ws_crc.append(
                [
                    hostname,
                    intf.get("interface", ""),
                    crc,
                    intf.get("link_status", ""),
                    intf.get("protocol_status", ""),
                    intf.get("description", ""),
                ]
            )

            if first_row_for_host is None:
                first_row_for_host = current_row

            current_row += 1
            rows_for_this_host += 1

        if first_row_for_host is not None and rows_for_this_host > 1:
            ws_crc.merge_cells(
                start_row=first_row_for_host,
                start_column=1,
                end_row=first_row_for_host + rows_for_this_host - 1,
                end_column=1,
            )

            master_cell = ws_crc.cell(first_row_for_host, 1)
            master_cell.value = hostname

            master_cell.alignment = Alignment(vertical="center")

    autosize_columns(ws_crc)

    # ---------------------------
    # Sheet 3: Logs (baru)
    # ---------------------------
    ws_logs: Worksheet = wb.create_sheet("Logs", 2)
    headers_logs = ["Hostname", "Timestamp", "Severity", "Message"]
    ws_logs.append(headers_logs)

    for hostname, device_data in data.items():
        logs = device_data.get("parsed_logs") or device_data.get("logs") or []
        if not isinstance(logs, list):
            continue

        for raw_line in logs:
            if raw_line is None:
                continue
            if not isinstance(raw_line, str):
                raw_line = str(raw_line)

            ts, severity, msg = _parse_syslog_line(raw_line)
            ws_logs.append([hostname, ts, severity, msg or raw_line])

    autosize_columns(ws_logs)

    # Bersihkan sheet default "Sheet" kalau masih ada dan kosong
    if "Sheet" in wb.sheetnames:
        std = wb["Sheet"]
        if std.max_row == 1 and std["A1"].value is None:
            wb.remove(std)

    wb.save(health_check_path)
    msg = f"Health-check saved to {health_check_path}"
    if progress_callback:
        try:
            progress_callback(msg)
        except Exception:
            pass
    else:
        print(msg)
    return health_check_path


def take_snapshot(base_dir=None, progress_callback=None):
    customer_name = get_customer_name()
    devices = load_devices()

    if base_dir:
        path = os.path.join(base_dir, customer_name, "legacy")
    else:
        path = os.path.join("results", customer_name, "legacy")

    os.makedirs(path, exist_ok=True)

    snapshot_dir = os.path.join(path, "snapshot")
    os.makedirs(snapshot_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    snapshot_path = os.path.join(
        snapshot_dir, f"{customer_name}_snapshot_{timestamp}.json"
    )

    result = {}
    for dev in devices:
        hostname = dev.get("hostname", "")
        data = capture_device_output(dev, progress_callback=progress_callback)
        result[hostname] = data

    with open(snapshot_path, "w") as f:
        json.dump(result, f, indent=2)

    msg = f"Snapshot saved to {snapshot_path}"
    if progress_callback:
        try:
            progress_callback(msg)
        except Exception:
            pass
    else:
        print(msg)

    health_path = health_check(
        customer_name, result, path, progress_callback=progress_callback
    )
    return snapshot_path, health_path