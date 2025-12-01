import os
import json
import re
import datetime
from deepdiff import DeepDiff
from rich.console import Console
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell

console = Console()


def _autosize_columns(ws, padding=2, min_width=12, max_width=80):
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            # Skip non-anchor merged cells
            if isinstance(cell, MergedCell):
                continue

            col_letter = get_column_letter(cell.column)
            val = "" if cell.value is None else str(cell.value)

            # If the cell has newlines, size to the longest line
            longest = max((len(line) for line in val.splitlines()), default=0)

            widths[col_letter] = max(widths.get(col_letter, 0), longest)

    for col, w in widths.items():
        ws.column_dimensions[col].width = max(min(w + padding, max_width), min_width)


def compare_snapshots(devices, file1, file2):
    hostnames = [dev["hostname"] for dev in devices]

    with (
        open(file1, "rt", encoding="utf-8") as f1,
        open(file2, "rt", encoding="utf-8") as f2,
    ):
        before_json = json.load(f1)
        after_json = json.load(f2)

    result = {}

    for host in hostnames:
        before = before_json.get(host, {})
        after = after_json.get(host, {})

        if not before or not after:
            continue

        result[host] = {}

        diff = DeepDiff(
            before,
            after,
            ignore_order=True,
            exclude_paths=["root['health_check']['uptime']"],
        )

        if not diff:
            console.print("[green]✅ No changes detected between snapshots.[/green]")
            return

        changed_values = diff.get("values_changed", {})
        added_items = diff.get("iterable_item_added", [])
        removed_items = diff.get("iterable_item_removed", [])

        result[host]["item_changes"] = {}
        result[host]["item_changes"]["interface_changes"] = []
        result[host]["item_changes"]["mac_address_table_changes"] = []
        result[host]["item_changes"]["routing_table_changes"] = []
        result[host]["item_changes"]["arp_table_changes"] = []
        result[host]["added_items"] = {}
        result[host]["added_items"]["mac_address_table_added"] = []
        result[host]["added_items"]["routing_table_added"] = []
        result[host]["added_items"]["arp_table_added"] = []
        result[host]["removed_items"] = {}
        result[host]["removed_items"]["mac_address_table_removed"] = []
        result[host]["removed_items"]["routing_table_removed"] = []
        result[host]["removed_items"]["arp_table_removed"] = []

        for path, change in changed_values.items():
            parts = [
                int(x) if x.isdigit() else x
                for x in re.findall(r"\['([^']+)'\]|\[(\d+)\]", path)
                for x in x
                if x
            ]
            if len(parts) >= 3:
                key, index, field = parts[0], parts[1], parts[2]

                if "interfaces" in path:
                    intf_details = before.get(key, {})[index]
                    result[host]["item_changes"]["interface_changes"].append(
                        {
                            "type": f"{field} changes",
                            "item": intf_details.get("interface", ""),
                            "before": change.get("old_value", ""),
                            "after": change.get("new_value", ""),
                        }
                    )

                elif "mac_address_table" in path:
                    mac_details = before.get(key, {})[index]
                    result[host]["item_changes"]["mac_address_table_changes"].append(
                        {
                            "type": f"{field} changes",
                            # "vlan": mac_details.get("vlan_id", ""),
                            "item": mac_details.get("mac_address", ""),
                            "before": change.get("old_value", ""),
                            "after": change.get("new_value", ""),
                        }
                    )

                elif "routing_table" in path:
                    routing_details = before.get(key, {})[index]
                    result[host]["item_changes"]["routing_table_changes"].append(
                        {
                            "type": f"{field} changes",
                            # "vrf": routing_details.get("vrf", ""),
                            # "protocol": routing_details.get("protocol", ""),
                            "item": routing_details.get("network", ""),
                            # "prefix_length": routing_details.get("prefix_length", ""),
                            # "nexthop_ip": routing_details.get("nexthop_ip", ""),
                            # "nexthop_if": routing_details.get("nexthop_if", ""),
                            "before": change.get("old_value", ""),
                            "after": change.get("new_value", ""),
                        }
                    )

                elif "arp_table" in path:
                    arp_details = before.get(key, {})[index]
                    result[host]["item_changes"]["arp_table_changes"].append(
                        {
                            "type": f"{field} changes",
                            # "vrf": arp_details.get("vrf", ""),
                            "item": arp_details.get("ip_address", ""),
                            # "mac_address": arp_details.get("mac_address", ""),
                            # "interface": arp_details.get("interface", ""),
                            "before": change.get("old_value", ""),
                            "after": change.get("new_value", ""),
                        }
                    )

        if added_items:
            for path, change in added_items.items():
                parts = [
                    int(x) if x.isdigit() else x
                    for x in re.findall(r"\['([^']+)'\]|\[(\d+)\]", path)
                    for x in x
                    if x
                ]

                key, index = parts[0], parts[1]

                if "mac_address_table" in path:
                    mac_details = after.get(key, {})[index]
                    result[host]["added_items"][f"{key}_added"].append(
                        {
                            "item": mac_details.get("mac_address", ""),
                            "details": f"{mac_details.get('vlan_id', '')}",
                        }
                    )

                elif "routing_table" in path:
                    routing_details = after.get(key, {})[index]
                    result[host]["added_items"][f"{key}_added"].append(
                        {
                            "item": f"{routing_details.get('network', '')}/{routing_details.get('prefix_length')}",
                            "details": f"Next hop IP: {routing_details.get('nexthop_ip', '')}",
                        }
                    )

                elif "arp_table" in path:
                    arp_details = after.get(key, {})[index]
                    result[host]["item_added"][f"{key}_added"].append(
                        {
                            "item": f"{arp_details.get('ip_address', '')} - {arp_details.get('mac_address', '')}",
                            "details": f"Port: {arp_details.get('ports', '')}",
                        }
                    )

        if removed_items:
            for path, change in removed_items.items():
                parts = [
                    int(x) if x.isdigit() else x
                    for x in re.findall(r"\['([^']+)'\]|\[(\d+)\]", path)
                    for x in x
                    if x
                ]

                key, index = parts[0], parts[1]

                if "mac_address_table" in path:
                    mac_details = before.get(key, {})[index]
                    result[host]["removed_items"][f"{key}_removed"].append(
                        {
                            "item": mac_details.get("mac_address", ""),
                            "details": mac_details.get("vlan_id", ""),
                        }
                    )

                elif "routing_table" in path:
                    routing_details = after.get(key, {})[index]
                    result[host]["removed_items"][f"{key}_removed"].append(
                        {
                            "item": f"{routing_details.get('network', '')}/{routing_details.get('prefix_length')}",
                            "details": f"Next hop IP: {routing_details.get('nexthop_ip', '')}",
                        }
                    )

                elif "arp_table" in path:
                    arp_details = after.get(key, {})[index]
                    result[host]["removed_items"][f"{key}_removed"].append(
                        {
                            "item": f"{arp_details.get('ip_address', '')} - {arp_details.get('mac_address', '')}",
                            "details": f"Port: {arp_details.get('ports', '')}",
                        }
                    )

    return result


def list_snapshots(dir):
    if not os.path.exists(dir):
        print("📂 No snapshots taken yet.")
        return []
    files = [f for f in os.listdir(dir) if f.endswith(".json")]
    if not files:
        print("📂 No snapshot files found.")
        return []
    files.sort()
    print("\n🕓 Available Snapshots:")
    for i, f in enumerate(files):
        print(f"  [{i + 1}] {f}")
    return files


def choose_snapshots(dir):
    files = list_snapshots(dir)
    if len(files) < 2:
        print("❌ Need at least 2 snapshots to compare.")
        return None, None
    try:
        first = int(input("🔢 Enter number for FIRST snapshot (before): ")) - 1
        second = int(input("🔢 Enter number for SECOND snapshot (after): ")) - 1
        if 0 <= first < len(files) and 0 <= second < len(files):
            return os.path.join(dir, files[first]), os.path.join(dir, files[second])
        else:
            print("❌ Invalid selection.")
            return None, None
    except ValueError:
        print("❌ Please enter valid numbers.")
        return None, None


def save_to_excel(
    all_results,
    devices,
    filepath,
):
    wb = Workbook()
    # remove default sheet
    wb.remove(wb.active)

    hostnames = [dev["hostname"] for dev in devices]

    for host in hostnames:
        result = all_results.get(host, {}) or {}
        if not result:
            continue
        ws = wb.create_sheet(host)

        ws.append(["Changes"])
        current_row = ws.max_row
        ws.merge_cells(
            start_row=current_row, start_column=1, end_row=current_row, end_column=5
        )
        current_cell = ws.cell(row=current_row, column=1)
        current_cell.alignment = Alignment(horizontal="center")

        ws.append(["Category", "Type", "Item", "Before", "After"])

        item_changes = result.get("item_changes", {})

        if item_changes:
            all_changes = []
            for category, changes in item_changes.items():
                for item in changes:
                    item["category"] = category.replace("_", " ")
                    all_changes.append(item)

            for item in all_changes:
                ws.append(
                    [
                        item.get("category", ""),
                        item.get("type", "").replace("_", " "),
                        item.get("item", ""),
                        item.get("before", ""),
                        item.get("after", ""),
                    ]
                )

        for _ in range(3):
            ws.append([])

        ws.append(["New Items"])
        current_row = ws.max_row
        ws.merge_cells(
            start_row=current_row, start_column=1, end_row=current_row, end_column=4
        )
        current_cell = ws.cell(row=current_row, column=1)
        current_cell.alignment = Alignment(horizontal="center")

        ws.append(["Category", "Item", "Detail"])

        items_added = result.get("added_items", {})

        if items_added:
            all_added = []
            for category, added in items_added.items():
                for item in added:
                    item["category"] = category.replace("_", " ")
                    all_added.append(item)

            for item in all_added:
                ws.append(
                    [
                        item.get("category", ""),
                        item.get("item", ""),
                        item.get("details", ""),
                    ]
                )

        for _ in range(3):
            ws.append([])

        ws.append(["Removed Items"])
        current_row = ws.max_row
        ws.merge_cells(
            start_row=current_row, start_column=1, end_row=current_row, end_column=4
        )
        current_cell = ws.cell(row=current_row, column=1)
        current_cell.alignment = Alignment(horizontal="center")

        ws.append(["Category", "Item", "Detail"])

        items_removed = result.get("removed_items", {})

        if items_removed:
            all_removed = []
            for category, removed in items_removed.items():
                for item in removed:
                    item["category"] = category.replace("_", " ")
                    all_removed.append(item)

            for item in all_removed:
                ws.append(
                    [
                        item.get("category", ""),
                        item.get("item", ""),
                        item.get("details", ""),
                    ]
                )

        _autosize_columns(ws)

    wb.save(filepath)


def compare(devices, base_dir: str | None = None):
    if base_dir:
        path = os.path.join(base_dir, "legacy", "compare")
        snapshot_path = os.path.join(base_dir, "legacy", "snapshot")
    else:
        path = os.path.join("legacy", "results", "compare")
        snapshot_path = os.path.join("legacy", "results", "snapshot")

    os.makedirs(path, exist_ok=True)

    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M")
    compare_path = os.path.join(path, f"comparison_result_{timestamp}.xlsx")

    print(snapshot_path)
    file1, file2 = choose_snapshots(snapshot_path)  # type: ignore
    if file1 and file2:
        print(f"📊 Comparing '{file1}' and '{file2}'...")
        results = compare_snapshots(devices, file1, file2)
        save_to_excel(results, devices, compare_path)
