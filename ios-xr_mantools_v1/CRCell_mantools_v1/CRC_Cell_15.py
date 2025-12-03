from netmiko import ConnectHandler
from netmiko import redispatch
import time
import re
import os
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from prettytable import PrettyTable

#### Login to router ####


banner = (f"""  
 ______________________________________________________________________
|:..                                                   ''':::::%%%%%%%%|
|%%%:::::..                      C R C E L L                 ''::::%%%%|
|%%%%%%%:::::.....__________________________________________________:::|
  
Description   : Automated CRC Error Monitoring with Delta Calculation
Version       : 1.0
------------------------------------------------------------------------
""")
print(banner)


#This section imports necessary libraries and defines the connection parameters for the jump server.
# Function to read username and password from a .txt file
def read_credentials(file_path):
    with open(file_path, 'r') as cred_file:
        lines = cred_file.readlines()
        username = lines[0].strip()
        password = lines[1].strip()
    return username, password

# Retrieve credentials
username, password = read_credentials('ssh_credentials.txt')

# Define jumpserver configuration
jumpserver = {
    "device_type": "terminal_server",
    "ip": "10.62.170.56",
    "username": username,
    "password": password,
    "port": 22
}
#---------------------------------------------------------------------------------------------------------------------

#---------------------------------------------------------------------------------------------------------------------
# Here, the script connects to the jump server and confirms the connection by printing the prompt.
# Connect to Jumphost
print(f"\n{'~'*50}\n1. Connecting to the Jump host\n{'~'*50}")
net_connect = ConnectHandler(**jumpserver)
print ("Jump Server Prompt: {}".format(net_connect.find_prompt()))
print(net_connect.find_prompt())
print(f"\n{'*'*10}Connected to the Jump host{'*'*10}")
#---------------------------------------------------------------------------------------------------------------------

#---------------------------------------------------------------------------------------------------------------------
# This section initiates an SSH connection to the destination router and reads the output.
# Connect to Router
print(f"\n{'~'*50}\n2. Connecting to the Dest Router\n{'~'*50}")
#net_connect.write_channel("telnet p-d2-boo\n")
destination = input("Enter the destination to connect (e.g: 'p-d2-bks', 'me-d6-sbr-4'): ")
connection_type = "telnet"
remote_node = f"{connection_type} {destination}"
net_connect.write_channel(f"{remote_node}\n")
time.sleep(3)
output = net_connect.read_channel()
print(output)
#---------------------------------------------------------------------------------------------------------------------

#---------------------------------------------------------------------------------------------------------------------
# If prompted for a password, the script sends the password and redispatches the connection to the appropriate device type.
if "username" in output:
    # Read username and password from a .txt file
    with open('tacacs_credentials.txt', 'r') as cred_file:
        lines = cred_file.readlines()
        router_username = lines[0].strip()
        router_password = lines[1].strip()
    
    # Send username and password to the router
    net_connect.write_channel(f"{router_username}\n")
    time.sleep(2)
    net_connect.write_channel(f"{router_password}\n")
    print(f"\n{'~'*50}\n3. Destination Device Prompt\n{'~'*50}\n")
    print(net_connect.find_prompt())
    print("Router Prompt: {}".format(net_connect.find_prompt()))
    redispatch(net_connect, device_type="cisco_ios")

# Print timestamp with day
#print(time.strftime("%a %d-%b-%Y %H:%M:%S WIB", time.localtime()))




####---------------------------------------------------------------------------------------------------------------------#####



# This section sends a command to the router and prints the output.
print(f"\n{'~'*50}\n5. Sending commands to the router\n{'~'*50}")

termlen_output = net_connect.send_command("terminal length 0")
print("Output for command 'term len 0':\n", termlen_output)

# Send command to show interfaces and filter out admin-down interfaces
show_int = net_connect.send_command("show interfaces hundredGigE * brief | exclude admin-down")
print("Output for command 'show interfaces hundredGigE * brief | exclude admin-down':\n", show_int)

# parse interface using regex
def parse_interface(data):
    return re.findall(r'\b(?:Hu)\S*\b', data)

# create command "show interface {interface} include CRC"
def generate_commands(parsed_interfaces):
    commands = []
    for interface in parsed_interfaces:
        command = f"show interface {interface} | include CRC"
        commands.append(command)
    return commands

# Generate commands for each interface
interfaces = parse_interface(show_int)
if not interfaces:
    print("WARNING: No HundredGigE interfaces were found. The script will terminate.")
    net_connect.disconnect()
    exit()

commands = generate_commands(interfaces)

# A dictionary to store all attempts, keyed by interface name
interface_attempts = {interface: [] for interface in interfaces}

# New Loop structure: Perform all attempts for each interface sequentially
# This section has been modified to handle the new looping logic
print(f"\n{'~'*50}\n5a. Executing commands for all interfaces (Attempt 1)\n{'~'*50}")
for command in commands:
    interface_name = re.search(r'\b(?:Hu)\S*\b', command).group(0)
    print(f"Executing command: '{command}' (Attempt 1)")
    show_crc = net_connect.send_command(command)
    
    # Parse the output and store
    input_match = re.search(r'(\d+)\s+input\s+errors', show_crc, re.IGNORECASE)
    crc_match = re.search(r'(\d+)\s+CRC', show_crc, re.IGNORECASE)
    output_match = re.search(r'(\d+)\s+output\s+errors', show_crc, re.IGNORECASE)

    input_errors = int(input_match.group(1)) if input_match else 0
    crc = int(crc_match.group(1)) if crc_match else 0
    output_errors = int(output_match.group(1)) if output_match else 0
    
    interface_attempts[interface_name].append({
        "input_errors": input_errors,
        "CRC": crc,
        "output_errors": output_errors
    })
    print(f"Command output (Attempt 1):\n{show_crc}\n")

# Delay before attempt 2
print("Waiting 2 second before starting Attempt 2...")
time.sleep(2)

print(f"\n{'~'*50}\n5b. Executing commands for all interfaces (Attempt 2)\n{'~'*50}")
for command in commands:
    interface_name = re.search(r'\b(?:Hu)\S*\b', command).group(0)
    print(f"Executing command: '{command}' (Attempt 2)")
    show_crc = net_connect.send_command(command)
    
    # Parse the output and store
    input_match = re.search(r'(\d+)\s+input\s+errors', show_crc, re.IGNORECASE)
    crc_match = re.search(r'(\d+)\s+CRC', show_crc, re.IGNORECASE)
    output_match = re.search(r'(\d+)\s+output\s+errors', show_crc, re.IGNORECASE)

    input_errors = int(input_match.group(1)) if input_match else 0
    crc = int(crc_match.group(1)) if crc_match else 0
    output_errors = int(output_match.group(1)) if output_match else 0
    
    interface_attempts[interface_name].append({
        "input_errors": input_errors,
        "CRC": crc,
        "output_errors": output_errors
    })
    print(f"Command output (Attempt 2):\n{show_crc}\n")

# Delay before attempt 3
print("Waiting 2 seconds before starting Attempt 3...")
time.sleep(2)

print(f"\n{'~'*50}\n5c. Executing commands for all interfaces (Attempt 3)\n{'~'*50}")
for command in commands:
    interface_name = re.search(r'\b(?:Hu)\S*\b', command).group(0)
    print(f"Executing command: '{command}' (Attempt 3)")
    show_crc = net_connect.send_command(command)
    
    # Parse the output and store
    input_match = re.search(r'(\d+)\s+input\s+errors', show_crc, re.IGNORECASE)
    crc_match = re.search(r'(\d+)\s+CRC', show_crc, re.IGNORECASE)
    output_match = re.search(r'(\d+)\s+output\s+errors', show_crc, re.IGNORECASE)

    input_errors = int(input_match.group(1)) if input_match else 0
    crc = int(crc_match.group(1)) if crc_match else 0
    output_errors = int(output_match.group(1)) if output_match else 0
    
    interface_attempts[interface_name].append({
        "input_errors": input_errors,
        "CRC": crc,
        "output_errors": output_errors
    })
    print(f"Command output (Attempt 3):\n{show_crc}\n")


# --- Data Extraction and Table Generation (Delta Calculation) ---
print(f"\n{'~'*50}\n6. Processing Data and Generating Table\n{'~'*50}")

# Prepare a list to hold the delta data
delta_data = []

for interface, attempts in interface_attempts.items():
    if len(attempts) >= 3:
        # Get data from attempt 1 and attempt 3
        attempt1 = attempts[0]
        attempt3 = attempts[2]
        
        # Calculate the delta (selisih) for each metric
        delta_input_errors = attempt3["input_errors"] - attempt1["input_errors"]
        delta_crc = attempt3["CRC"] - attempt1["CRC"]
        delta_output_errors = attempt3["output_errors"] - attempt1["output_errors"]
        
        delta_data.append({
            "interface": interface,
            "timestamp": datetime.now().strftime("%a %d-%b-%Y %H:%M:%S WIB"),
            "delta_input_errors": delta_input_errors,
            "delta_crc": delta_crc,
            "delta_output_errors": delta_output_errors
        })
    else:
        print(f"WARNING: Insufficient attempts for interface {interface}. Skipped delta calculation.")

# Create a PrettyTable to display the delta data
table = PrettyTable()

# Update field names to reflect the delta values
table.field_names = ["Interface", "Timestamp", "Delta Input Errors", "Delta CRC", "Delta Output Errors"]

for entry in delta_data:
    table.add_row([
        entry["interface"],
        entry["timestamp"],
        entry["delta_input_errors"],
        entry["delta_crc"],
        entry["delta_output_errors"]
    ])
    
print(table)

# --- File Saving .txt (Dipertahankan untuk cadangan) ---
print(f"\n{'~'*50}\n7. Writing output to file\n{'~'*50}")
timestamp = datetime.now().strftime("%d-%b-%Y_%H-%M-%S_WIB")
txt_filename = f'CRC_DELTA_{timestamp}.txt'
with open(txt_filename, 'w') as f:
    f.write(str(table))
print(f"Data saved to text file: {txt_filename}")


# --- File Saving .xlsx (Menyimpan, meng-append, dan mewarnai) ---
excel_filename = 'CRC_DELTA_SUMMARY.xlsx'
sheet_name = destination
red_font = Font(color="FF0000") # Definisi font merah

try:
    df_new = pd.DataFrame(delta_data)
    sheets_data = {}
    
    if os.path.exists(excel_filename):
        sheets_data = pd.read_excel(excel_filename, sheet_name=None)
        if sheet_name in sheets_data:
            df_old = sheets_data[sheet_name]
            df_combined = pd.concat([df_old, df_new], ignore_index=True)
            sheets_data[sheet_name] = df_combined
        else:
            sheets_data[sheet_name] = df_new
    else:
        sheets_data[sheet_name] = df_new

    # Menulis semua data ke Excel
    with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
        for sheet_title, df_content in sheets_data.items():
            df_content.to_excel(writer, sheet_name=sheet_title, index=False)
    
    # --- Tambahan untuk Pewarnaan Font ---
    print(f"Applying conditional formatting to sheet '{sheet_name}'...")
    wb = load_workbook(excel_filename)
    ws = wb[sheet_name]

    # Cari nomor kolom untuk setiap header
    col_map = {header: idx + 1 for idx, header in enumerate(ws[1])}
    
    col_input = col_map.get("delta_input_errors")
    col_crc = col_map.get("delta_crc")
    col_output = col_map.get("delta_output_errors")

    # Iterasi dari baris kedua (data, bukan header) hingga akhir
    for row in range(2, ws.max_row + 1):
        if col_input and ws.cell(row=row, column=col_input).value > 0:
            ws.cell(row=row, column=col_input).font = red_font
        if col_crc and ws.cell(row=row, column=col_crc).value > 0:
            ws.cell(row=row, column=col_crc).font = red_font
        if col_output and ws.cell(row=row, column=col_output).value > 0:
            ws.cell(row=row, column=col_output).font = red_font
    
    # Menyimpan file Excel dengan format yang sudah diterapkan
    wb.save(excel_filename)
    
    print(f"Data saved and formatted in Excel file: {excel_filename} on sheet '{sheet_name}'")

except ImportError:
    print("Warning: pandas or openpyxl not installed. Could not save to Excel.")


# Disconnect from the device
print(f"\n{'~'*50}\n8. Disconnecting from device\n{'~'*50}")
net_connect.disconnect()