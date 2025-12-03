import os
from netmiko import ConnectHandler
from netmiko import redispatch
import time
import re
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
#### Login to router ####


banner = (f"""
    
 _________________________________________________________________________________________
|:..                                                                     ''':::::%%%%%%%%|
|%%%:::::..                               X R A Y                              ''::::%%%%|
|%%%%%%%:::::.....____________________________________________________________________:::|
  
Description   : Automated IOS-XR dBm optic 100G interface monitoring with Excel reporting
Version       : 1.0
-----------------------------------------------------------------------------------------

""")
print(banner)


#This section imports necessary libraries and defines the connection parameters for the jump server.
# Function to read username and password from a .txt file
def read_credentials(file_path):
    with open(file_path, 'r') as cred_file:
        lines = cred_file.readlines()
        username = lines[0].strip()
        password = lines[1].strip()
    return username, password

# Retrieve credentials
username, password = read_credentials('ssh_credentials.txt')

# Define jumpserver configuration
jumpserver = {
    "device_type": "terminal_server",
    "ip": "10.62.170.56",
    "username": username,
    "password": password,
    "port": 22
}
#---------------------------------------------------------------------------------------------------------------------

#---------------------------------------------------------------------------------------------------------------------
# Here, the script connects to the jump server and confirms the connection by printing the prompt.
# Connect to Jumphost
print(f"\n{'~'*50}\n1. Connecting to the Jump host\n{'~'*50}")
net_connect = ConnectHandler(**jumpserver)
print ("Jump Server Prompt: {}".format(net_connect.find_prompt()))
print(net_connect.find_prompt())
print(f"\n{'*'*10}Connected to the Jump host{'*'*10}")
#---------------------------------------------------------------------------------------------------------------------

#---------------------------------------------------------------------------------------------------------------------
# This section initiates an SSH connection to the destination router and reads the output.
# Connect to Router
print(f"\n{'~'*50}\n2. Connecting to the Dest Router\n{'~'*50}")
#net_connect.write_channel("telnet p-d2-boo\n")
destination = input("Enter the destination to connect (e.g: 'p-d2-bks', 'me-d6-sbr-4'): ")
connection_type = "telnet"
remote_node = f"{connection_type} {destination}"
net_connect.write_channel(f"{remote_node}\n")
time.sleep(3)
output = net_connect.read_channel()
print(output)
#---------------------------------------------------------------------------------------------------------------------

#---------------------------------------------------------------------------------------------------------------------
# If prompted for a password, the script sends the password and redispatches the connection to the appropriate device type.
if "username" in output:
    # Read username and password from a .txt file
    with open('tacacs_credentials.txt', 'r') as cred_file:
        lines = cred_file.readlines()
        router_username = lines[0].strip()
        router_password = lines[1].strip()
    
    # Send username and password to the router
    net_connect.write_channel(f"{router_username}\n")
    time.sleep(2)
    net_connect.write_channel(f"{router_password}\n")
    print(f"\n{'~'*50}\n3. Destination Device Prompt\n{'~'*50}\n")
    print(net_connect.find_prompt())
    print("Router Prompt: {}".format(net_connect.find_prompt()))
    redispatch(net_connect, device_type="cisco_ios")

# Print timestamp with day
#print(time.strftime("%a %d-%b-%Y %H:%M:%S WIB", time.localtime()))


####---------------------------------------------------------------------------------------------------------------------

# This section sends a command to the router and prints the output.
print(f"\n{'~'*50}\n4. Sending command to the router\n{'~'*50}")

# Send command in router
# terleng = net_connect.send_command("terminal length 0")
# print("Output for command 'term len 0':\n", terleng)
show_int = net_connect.send_command("show interfaces hundredGigE * brief | exclude down")
print("Output for command 'show interfaces hundredGigE * brief | exclude down':\n", show_int)

# parse interface using regex
def parse_interface(data):
    parse_interface = re.findall(r'\b(?:Hu)\S*\b', data)
    return parse_interface

# create command "show controllers {interface} phy | in dB | in Amp"
def generate_commands(parse_interface):
    commands = []
    for interface in parse_interface:
        command = f"show controllers {interface} phy | in dB | in Amp"
        commands.append(command)
    return commands

# Eksekusi show controllers {interface} phy | in dB | in Amp
commands = generate_commands(parse_interface(show_int))
show_controllers_list = []
for command in commands:
    show_controllers = net_connect.send_command(command)
    show_controllers_list.append(f"#################################'{command}':#################################\n{show_controllers}\n")
    print(f"Output for command '{command}':\n{show_controllers}\n")

# Timestamp for filename
timestamp = datetime.now().strftime('%d-%m-%Y_%H-%M-%S')
output_file = f"txrx_{destination}_{timestamp}.txt"

# Write and save the data to file
txrx_output = "\n\n".join(show_controllers_list)
with open(output_file, 'w') as outfile:
    outfile.write(txrx_output)
print(f'Successfully saved to {output_file}.')

# Disconnect from device
net_connect.disconnect()

# Function to extract data from file
def read_raw_data(file_path):
    with open(file_path, 'r') as file:
        return file.read()

# Function to extract data from raw data
def extract_data(raw_data, destination): # Add 'destination' as an argument
    data = []
    lines = raw_data.splitlines()
    port = None

    # Tambahkan timestamp di sini
    current_timestamp = datetime.now().strftime("%a %d-%b-%Y %H:%M:%S WIB")

    for line in lines:
        # Extract port and date
        port_match = re.match(r'.*show controllers (Hu\S*)', line)
        if port_match:
            port = port_match.group(1)

        # Extract lane, Tx Power, and Rx Power
        value_match = re.match(r'\s+(\d)\s+.*\s+\d+\.\d+ mAmps\s+\d+\.\d+ mW \(([-\d.]+ dBm)\)\s+\d+\.\d+ mW \(([-\d.]+ dBm)\)', line)
        if value_match:
            lane = value_match.group(1)
            tx_power = value_match.group(2)
            rx_power = value_match.group(3)
            if port and lane and tx_power and rx_power:
                # Masukkan timestamp ke dalam list data
                data.append([destination, current_timestamp, port, lane, tx_power, rx_power]) 

    return data

# Main function
file_path = output_file  # Use the dynamically generated file name
raw_data = read_raw_data(file_path)

# Extracted data
extracted_data = extract_data(raw_data, destination) # Pass 'destination' to the function

# --- File Saving .xlsx (Bagian ini yang diubah total) ---
print(f"\n{'~'*50}\n7. Writing output to file\n{'~'*50}")

excel_filename = 'xRay_dBm_interface.xlsx'
sheet_name = destination

try:
    # Buat DataFrame dari data baru dengan kolom timestamp
    df_new = pd.DataFrame(extracted_data, columns=["Hostname", "Timestamp", "Port", "Lane", "Tx Power", "Rx Power"])
    
    # Inisialisasi dictionary untuk menyimpan semua data sheet
    sheets_data = {}
    
    # Cek apakah file Excel sudah ada
    if os.path.exists(excel_filename):
        # Jika file ada, baca semua sheet yang ada
        sheets_data = pd.read_excel(excel_filename, sheet_name=None)
        
        # Cek apakah sheet untuk destinasi ini sudah ada
        if sheet_name in sheets_data:
            # Jika sheet sudah ada, gabungkan data lama dengan data baru
            df_old = sheets_data[sheet_name]
            df_combined = pd.concat([df_old, df_new], ignore_index=True)
            sheets_data[sheet_name] = df_combined
        else:
            # Jika sheet belum ada, tambahkan data baru sebagai sheet baru
            sheets_data[sheet_name] = df_new
    else:
        # Jika file belum ada, buat sheet baru dengan data baru
        sheets_data[sheet_name] = df_new

    # Simpan semua data sheet ke file Excel menggunakan ExcelWriter
    with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
        for sheet_title, df_content in sheets_data.items():
            df_content.to_excel(writer, sheet_name=sheet_title, index=False)
            
    print(f"Data saved and appended to Excel file: {excel_filename} on sheet '{sheet_name}'")

    # --- Bagian Openpyxl untuk Pemformatan (Ditambahkan kembali di sini) ---
    # Load the workbook and select the active sheet
    wb = load_workbook(excel_filename)
    ws = wb[sheet_name]

    # Set column widths for A, B, C, and D
    columns_to_adjust = ['A', 'B', 'C', 'D', 'E']
    for col in columns_to_adjust:
        ws.column_dimensions[col].width = 30

    # Define the starting column for the legend
    legend_column = len(df_new.columns) + 2

    # Add the signature and legend to the far right
    ws.cell(row=1, column=legend_column, value="***this file is converted by wingman-x/x-ray.py")
    ws.cell(row=3, column=legend_column, value="Thresholds:")
    ws.cell(row=4, column=legend_column, value="Red Font: Tx-Rx power Alarm high (7.499)")
    ws.cell(row=5, column=legend_column, value="Yellow Fill: Tx power Alarm low (-8.3)")
    ws.cell(row=6, column=legend_column, value="Red Fill: Rx power Alarm low  (-14.5)")

    # Define the highlighting rules
    red_font = Font(color="FF0000")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    for row in ws.iter_rows(min_row=2, min_col=4, max_col=5):  # Adjust to Tx Power and Rx Power columns
        for cell in row:
            value = float(str(cell.value).replace(" dBm", ""))
            if value >= 7.499:
                cell.font = red_font
            if cell.column_letter == "D" and value <= -8.3:
                cell.fill = yellow_fill
            if cell.column_letter == "E" and value <= -14.5:
                cell.fill = red_fill

    # Define the bottom border style
    bottom_border = Border(bottom=Side(style='thin'))

    # Apply bottom border to rows where Lane is 3
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=5):
        if str(row[3].value) == '3':  # Checking the Lane column (D)
            for cell in row:
                cell.border = bottom_border

    # Save the workbook
    wb.save(excel_filename)
    print(f"Excel file formatted and saved as {excel_filename}.")


except ImportError:
    print("Warning: pandas or openpyxl not installed. Could not save to Excel.")
except Exception as e:
    print(f"An error occurred while saving or formatting the Excel file: {e}")

# --- File Saving .txt (Bagian ini bisa Anda pertahankan atau hapus) ---
# ... (Kode untuk menyimpan ke file .txt tetap sama) ...

# Disconnect from the device
print(f"\n{'~'*50}\n8. Disconnecting from device\n{'~'*50}")
net_connect.disconnect()