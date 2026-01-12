import os
import csv
import requests
from typing import Optional
from requests.cookies import RequestsCookieJar
from cryptography.fernet import Fernet
from inventory.lib.credential_manager import load_key
from rich.console import Console
from rich.table import Column
from rich import print as rprint
from legacy.customer_context import get_customer_name
import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
import re

DEBUG = True

def debug(msg):
    if DEBUG:
        console.print(f"[dim][DEBUG][/dim] {msg}")

console = Console()

from inventory.lib.path import inventory_path
def load_devices(file=None):
    file = inventory_path() if file is None else file
    devices = []
    try:
        with open(file, "r") as f:
            reader = csv.reader(f, delimiter=";")

            for row in reader:
                if len(row) != 5:
                    continue

                hostname, ip, os_type, username, enc_password = row

                if "apic" not in os_type:
                    continue

                devices.append(
                    {
                        "hostname": hostname,
                        "ip": ip,
                        "os": os_type,
                        "username": username,
                        "password": enc_password,
                    }
                )

        return devices

    except FileNotFoundError:
        return []


def apic_login(
    apic_ip: str, username: str, password: str
) -> Optional[RequestsCookieJar]:
    key = load_key()
    fernet = Fernet(key)

    login_url = f"https://{apic_ip}/api/aaaLogin.json"
    auth_payload = {
        "aaaUser": {
            "attributes": {
                "name": username,
                "pwd": fernet.decrypt(password.encode()).decode(),
            }
        }
    }

    try:
        resp = requests.post(login_url, json=auth_payload, verify=False, timeout=30)
        if resp.status_code != 200:
            print(f"✗ Login failed with status code: {resp.status_code}")
            return None

        data = resp.json()
        if "imdata" in data and len(data["imdata"]) > 0:
            if isinstance(data["imdata"][0], dict) and "error" in data["imdata"][0]:
                print("✗ Authentication failed: Invalid credentials.")
                return None

        console.print(f"[green]✓ Successfully authenticated to APIC {apic_ip}[/green]")
        return resp.cookies

    except requests.exceptions.ConnectionError:
        print(f"✗ Cannot connect to APIC at {apic_ip}")
    except requests.exceptions.Timeout:
        print("✗ Connection timeout.")
    except Exception as e:
        print(f"✗ Login failed: {str(e)}")

from rich.table import Table
from rich.console import Console

console = Console()

from rich.table import Table

def print_general_table(apic, result):
    table = Table(
        title=f"[bold cyan]{apic}[/bold cyan] - [bold yellow]General[/bold yellow]",
        show_header=True,
        header_style="bold cyan",
    )

    table.add_column("Category")
    table.add_column("Item")
    table.add_column("Details")

    # =====================
    # Fabric Health
    # =====================
    fh = result.get("fabric_health", {})
    table.add_row("Fabric Health", "Before", str(fh.get("before", "N/A")))
    table.add_row("Fabric Health", "After", str(fh.get("after", "N/A")))

    console.print(table)

def parse_status_change(line):
    try:
        intf, node, rest = line.split("|", 2)
        before, after = rest.split("➜")
        return intf.strip(), node.strip(), before.strip(), after.strip()
    except ValueError:
        return None, None, None, None


def print_interface_errors_table(apic, result):
    table = Table(
        title=f"[bold cyan]{apic}[/bold cyan] - [bold yellow]Interface Errors[/bold yellow]",
        show_header=True,
        header_style="bold cyan",
    )

    headers = [
        "Category",
        "Node",
        "Port",
        "Port Description",
        "Before",
        "After",
        "Mac",
        "IP",
        "VLAN",
        "EPG Description",
    ]

    for h in headers:
        table.add_column(h)

    def add_rows(category, rows):
        for row in rows or []:
            endpoints = row.get("endpoints", [])
            if endpoints:
                for ep in endpoints:
                    table.add_row(
                        category,
                        str(row.get("node", "")),
                        str(row.get("interface", "")),
                        str(row.get("descr", "")),
                        str(row.get("before", "")),
                        str(row.get("after", "")),
                        str(ep.get("mac", "")),
                        str(ep.get("ip", "")),
                        str(ep.get("vlan", "")),
                        str(ep.get("epg_descr", "")),
                    )
            else:
                table.add_row(
                    category,
                    str(row.get("node", "")),
                    str(row.get("interface", "")),
                    str(row.get("descr", "")),
                    str(row.get("before", "")),
                    str(row.get("after", "")),
                    str(row.get("mac", "")),
                    str(row.get("ip", "")),
                    str(row.get("vlan", "")),
                    str(row.get("epg_descr", "")),
                )

    # =====================
    # Interface Status Changes
    # =====================
    intf_changes = result.get("interface_changes", {}) or {}

    for change in intf_changes.get("status_changed", []):
        intf, node, before, after = parse_status_change(change)
        if not intf:
            continue

        table.add_row(
            "Interface Status",
            node,
            intf,
            "",
            before,
            after,
            "",
            "",
            "",
            "",
        )



    add_rows("CRC Changes", result.get("crc_error_changes", []))
    add_rows("Drop Changes", result.get("drop_error_changes", []))
    add_rows("Output Error Changes", result.get("output_error_changes", []))

    console.print(table)

def print_endpoints_table(apic, result):
    table = Table(
        title=f"[bold cyan]{apic}[/bold cyan] - [bold yellow]Endpoints[/bold yellow]",
        show_header=True,
        header_style="bold cyan",
    )

    headers = [
        "Category",
        "Endpoint",
        "Mac",
        "IP address",
        "Switch",
        "Interface",
        "VLAN",
        "Description",
    ]

    for h in headers:
        table.add_column(h)

    for cat, key in (
        ("New Endpoints", "new_endpoints"),
        ("Missing Endpoints", "missing_endpoints"),
    ):
        for ep in result.get(key, []) or []:
            if not isinstance(ep, dict):
                continue
            table.add_row(
                cat,
                ep.get("dn", ""),
                ep.get("mac", ""),
                ep.get("ip", ""),
                ep.get("node", ""),
                ep.get("interface", ""),
                ep.get("vlan", ""),
                ep.get("epg_descr", ""),
            )

    console.print(table)

def normalize_faults(faults):
    out = {}
    for f in faults:
        attr = f.get("faultInst", {}).get("attributes", {})
        dn = attr.get("dn")
        if not dn:
            continue

        out[dn] = {
            "dn": dn,
            "code": attr.get("code", ""),
            "created": attr.get("created", ""),
            "severity": attr.get("severity", ""),
            "descr": attr.get("descr", ""),
        }
    return out

def print_faults_table(apic, result):
    table = Table(
        title=f"[bold cyan]{apic}[/bold cyan] - [bold yellow]Faults[/bold yellow]",
        show_header=True,
        header_style="bold cyan",
    )

    table.add_column("Category", style="bold")
    table.add_column("Code", no_wrap=True)
    table.add_column("Severity", no_wrap=True)
    table.add_column(
        "DN",
        overflow="fold",
        max_width=80,
    )
    table.add_column(
        "Description",
        overflow="fold",
        max_width=80,
    )
    table.add_column("Created", no_wrap=True)

    # ---------------------
    # New Faults
    # ---------------------
    for fault in result.get("new_faults", []):
        table.add_row(
            "New Fault",
            fault.get("code", ""),
            fault.get("severity", ""),
            fault.get("dn", ""),
            fault.get("descr", ""),
            fault.get("created", ""),
        )

    # ---------------------
    # Cleared Faults
    # ---------------------
    for fault in result.get("cleared_faults", []):
        table.add_row(
            "Cleared Fault",
            fault.get("code", ""),
            fault.get("severity", ""),
            fault.get("dn", ""),
            fault.get("descr", ""),
            fault.get("created", ""),
        )

    if table.row_count > 0:
        console.print(table)

def parse_urib_dn(dn):
    """
    Example DN:
    topology/pod-1/node-201/sys/uribv4/dom-overlay-1/db-rt/rt-[10.0.152.67/32]
    """
    node = ""
    domain = ""
    prefix = ""

    parts = dn.split("/")

    for p in parts:
        if p.startswith("node-"):
            node = p
        elif p.startswith("dom-"):
            domain = p.replace("dom-", "")

    if "rt-[" in dn:
        prefix = dn.split("rt-[", 1)[1].rstrip("]")

    return node, domain, prefix

def print_urib_routes_table(apic, result):
    table = Table(
        title=f"[bold cyan]{apic}[/bold cyan] - [bold yellow]URIB Routes[/bold yellow]",
        show_header=True,
        header_style="bold cyan",
    )

    table.add_column("Category", style="bold")
    table.add_column("Node", no_wrap=True)
    table.add_column("Domain / VRF", no_wrap=True)
    table.add_column("Prefix", no_wrap=True)
    table.add_column("DN", overflow="fold", max_width=90)

    urib = result.get("urib_route_changes", {}) or {}

    # =====================
    # New Routes
    # =====================
    for dn in urib.get("new", []):
        node, domain, prefix = parse_urib_dn(dn)
        table.add_row(
            "New Route",
            node,
            domain,
            prefix,
            dn,
        )

    # =====================
    # Missing Routes
    # =====================
    for dn in urib.get("missing", []):
        node, domain, prefix = parse_urib_dn(dn)
        table.add_row(
            "Missing Route",
            node,
            domain,
            prefix,
            dn,
        )

    if table.row_count:
        console.print(table)

def print_colored_result(result):
    rprint("\n📈 [bold]COMPARISON RESULT:[/bold]\n")

    if not result:
        rprint("[dim](no differences found)[/dim]")
        return

    for apic, apic_result in result.items():
        rprint(f"\n🏷️ [bold magenta]APIC: {apic}[/bold magenta]\n")

        print_general_table(apic, apic_result)
        print("\n")
        print_interface_errors_table(apic, apic_result)
        print("\n")
        print_endpoints_table(apic, apic_result) 
        print("\n")      
        print_faults_table(apic, apic_result)
        print("\n") 
        print_urib_routes_table(apic, apic_result)

def _autosize_columns(ws, padding=2, min_width=12, max_width=80):
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            # Skip non-anchor merged cells
            if isinstance(cell, MergedCell):
                continue

            col_letter = get_column_letter(cell.column)
            val = "" if cell.value is None else str(cell.value)

            # If the cell has newlines, size to the longest line
            longest = max((len(line) for line in val.splitlines()), default=0)

            widths[col_letter] = max(widths.get(col_letter, 0), longest)

    for col, w in widths.items():
        ws.column_dimensions[col].width = max(min(w + padding, max_width), min_width)

def sanitize_sheet_title(title: str, max_len: int = 31) -> str:
    """Return an Excel-safe sheet title (valid chars + max length)."""
    # Excel sheet Guard excluding sensitive char.
    invalid_chars = set(':/\\?*[]')
    cleaned = "".join(ch for ch in title if ch not in invalid_chars)
    return cleaned[:max_len].strip() or "Sheet"

def unique_sheet_title(base: str, used: set, max_len: int = 31) -> str:
    """Return a unique, safe sheet title within the workbook."""
    title = sanitize_sheet_title(base, max_len=max_len)
    if title not in used:
        used.add(title)
        return title

    # Add a suffix while keeping within Excel's 31-char limit.
    suffix = 1
    while True:
        suffix_str = f"~{suffix}"
        trimmed = title[: max_len - len(suffix_str)]
        candidate = f"{trimmed}{suffix_str}"
        if candidate not in used:
            used.add(candidate)
            return candidate
        suffix += 1

def save_to_excel(all_result: dict, filename=None, base_dir=None):
    apics = list(all_result.keys())
    customer_name = get_customer_name()

    # Create directory structure
    if base_dir:
        compare_dir = os.path.join(base_dir, customer_name, "aci", "compare")
    else:
        compare_dir = os.path.join("results", customer_name, "aci", "compare")
    os.makedirs(compare_dir, exist_ok=True)

    if filename is None:
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        filename = f"{customer_name}_comparison_result_{timestamp}.xlsx"

    filepath = os.path.join(compare_dir, filename)

    wb = Workbook()
    wb.remove(wb.active)  # type: ignore

    debug(f"save_to_excel: APICs to write = {apics}")
    used_titles = set()
    for apic in apics:
        result = all_result.get(apic, {}) or {}
        ws = wb.create_sheet(unique_sheet_title(f"{apic} - general", used_titles))

        ws.append(["Category", "Item", "Details"])

        fh = result.get("fabric_health", {})
        ws.append(["Fabric Health", "Before", fh.get("before", "N/A")])
        ws.append(["Fabric Health", "After", fh.get("after", "N/A")])
        _autosize_columns(ws)

        # NOTE: Faults Changes
        ws = wb.create_sheet(unique_sheet_title(f"{apic} - Faults Changes", used_titles))

        headers = [
            "Category",
            "Code",
            "Severity",
            "DN",
            "Description",
            "Created"
        ]
        ws.append(headers)

        # New Faults
        for fault in result.get("new_faults", []):
            if not isinstance(fault, dict):
                continue
            ws.append([
                "New Fault",
                fault.get("code", ""),
                fault.get("severity", ""),
                fault.get("dn", ""),
                fault.get("descr", ""),
                fault.get("created", ""),
            ])

        # Cleared Faults
        for fault in result.get("cleared_faults", []):
            if not isinstance(fault, dict):
                continue
            ws.append([
                "Cleared Fault",
                fault.get("code", ""),
                fault.get("severity", ""),
                fault.get("dn", ""),
                fault.get("descr", ""),
                fault.get("created", ""),
            ])

        # Wrap DN + Description
        for row in ws.iter_rows(min_row=2):
            row[3].alignment = Alignment(wrap_text=True)
            row[4].alignment = Alignment(wrap_text=True)
            ws.row_dimensions[row[0].row].height = 60
        _autosize_columns(ws)

        # NOTE: Route Changes
        ws = wb.create_sheet(unique_sheet_title(f"{apic} - Route Changes", used_titles))

        headers = [
            "Category",
            "Node",
            "Domain / VRF",
            "Prefix",
            "DN",
        ]
        ws.append(headers)

        urib = result.get("urib_route_changes", {}) or {}

        # New Routes
        for dn in urib.get("new", []):
            node, domain, prefix = parse_urib_dn(dn)
            ws.append([
                "New Route",
                node,
                domain,
                prefix,
                dn,
            ])

        # Missing Routes
        for dn in urib.get("missing", []):
            node, domain, prefix = parse_urib_dn(dn)
            ws.append([
                "Missing Route",
                node,
                domain,
                prefix,
                dn,
            ])

        _autosize_columns(ws)

        # NOTE: Interface Error Sheet
        int_sheet_name = f"{apic} - Interface Errors"
        intf_ws = wb.create_sheet(int_sheet_name[:31])

        headers = [
            "Category",
            "Node",
            "Interface",
            "Description",
            "Before",
            "After",
            "Mac",
            "IP",
            "VLAN",
            "Description",
        ]
        intf_ws.append(headers)

        cec = result.get("crc_error_changes", [])
        dec = result.get("drop_error_changes", [])
        oec = result.get("output_error_changes", [])

        write_interface_errors(intf_ws, "CRC Changes", cec)
        write_interface_errors(intf_ws, "Drop Changes", dec)
        write_interface_errors(intf_ws, "Output Error Changes", oec)

        # NOTE: Endpoints Sheet
        eps_sheet_name = f"{apic} - Endpoints"
        eps_ws = wb.create_sheet(eps_sheet_name[:31])

        eps_ws.append(
            [
                "Category",
                "Endpoint",
                "Mac",
                "Ip address",
                "Switch",
                "interface",
                "vlan",
                "Description",
            ]
        )

        for cat, key in (
            ("New Endpoints", "new_endpoints"),
            ("Missing Endpoints", "missing_endpoints"),
        ):
            for ep in result.get(key, []) or []:
                if not isinstance(ep, dict):
                    continue
                eps_ws.append(
                    [
                        cat,
                        ep.get("dn", ""),
                        ep.get("mac", ""),
                        ep.get("ip", ""),
                        ep.get("node", ""),
                        ep.get("interface", ""),
                        ep.get("vlan", ""),
                        ep.get("epg_descr", ""),
                    ]
                )

            _autosize_columns(eps_ws)

        wb.save(filepath)
    console.print(f"[cyan] ✓ Saved comparison result to: {filepath}[/cyan]")

def summarize_ethpm_interfaces(data):
    result = {}
    for item in data:
        attrs = item.get("ethpmPhysIf", {}).get("attributes", {})
        dn = attrs.get("dn", "")
        oper = attrs.get("operSt")
        qual = attrs.get("operStQual")

        # Extract eth1/33 from DN
        if "phys-[" in dn:
            intf = dn.split("phys-[")[-1].split("]")[0]
            if oper:
                result[intf] = f"{oper} ({qual})" if qual else oper

    debug(f"summarize_ethpm_interfaces: parsed={len(result)}")
    return result


def summarize_interfaces(l1_data, ethpm_data):
    l1_map = {}
    for item in l1_data:
        attrs = item.get("l1PhysIf", {}).get("attributes", {})
        dn = attrs.get("dn", "")
        intf = attrs.get("id")

        # Extract node from DN
        node = ""
        for part in dn.split("/"):
            if part.startswith("node-"):
                node = part
                break
        if intf:
            l1_map[intf] = {
                "node": node,
                "cfg": attrs.get("switchingSt", "unknown"),
            }

    ethpm_map = summarize_ethpm_interfaces(ethpm_data)

    merged = {}
    for intf, info in l1_map.items():
        oper_state = ethpm_map.get(intf, "unknown")
        merged[intf] = {
            "node": info["node"],
            "status": f"{info['cfg']} / {oper_state}",
        }

    debug(f"summarize_interfaces: merged interfaces={len(merged)}")
    return merged

def summarize_interface_errors(interface_errors):
    summary = {}
    for entry in interface_errors:
        dn = entry.get("dn")
        crc = int(entry.get("crc", 0))
        input_discards = int(entry.get("inputDiscards", 0))
        state = int(entry.get("operSt", 0))
        total_errors = crc + input_discards
        if dn:
            summary[dn] = total_errors
    return summary


def extract_interface_from_dn(dn):
    """
    Extract node ID and port from DN string.
    Example input: "topology/pod-1/node-102/sys/phys-[eth1/5]/dbgEtherStats"
    Output: ("node-102", "eth1/5")
    """
    match = re.search(r"node-(\d+).*phys-\[(.*?)\]", dn)
    if match:
        node_id = f"node-{match.group(1)}"
        port = match.group(2)
        return node_id, port
    return None, None


def write_interface_errors(intf_ws, category: str, changes: list):
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for entry in changes:
        node = entry["node"]
        interface = entry["interface"]
        intf_descr = entry["interface_descr"]
        before = entry["before"]
        after = entry["after"]
        endpoints = entry.get("endpoints", [])

        start_row = intf_ws.max_row + 1

        if endpoints:
            for ep in endpoints:
                intf_ws.append(
                    [
                        category,
                        node,
                        interface,
                        intf_descr,
                        before,
                        after,
                        ep.get("mac", ""),
                        ep.get("ip", ""),
                        ep.get("vlan", ""),
                        ep.get("epg_descr", ""),
                    ]
                )
            end_row = intf_ws.max_row

            for col in ("A", "B", "C", "D", "E", "F"):
                intf_ws.merge_cells(f"{col}{start_row}:{col}{end_row}")
                intf_ws[f"{col}{start_row}"].alignment = align_center

        else:
            intf_ws.append(
                [category, node, interface, intf_descr, before, after, "", "", "", ""]
            )
            r = intf_ws.max_row
            # Align the entire single row
            for col in ("A", "B", "C", "D", "E", "F", "G", "H"):
                intf_ws[f"{col}{r}"].alignment = align_center

    _autosize_columns(intf_ws)
