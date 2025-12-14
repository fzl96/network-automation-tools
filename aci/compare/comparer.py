import os
import json
import re
import datetime
import glob
from rich import print as rprint
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
from aci.snapshot.snapshotter import choose_snapshots
from aci.lib.utils import load_devices
from legacy.customer_context import get_customer_name

def _autosize_columns(ws, padding=2, min_width=12, max_width=80):
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            # Skip non-anchor merged cells
            if isinstance(cell, MergedCell):
                continue

            col_letter = get_column_letter(cell.column)
            val = "" if cell.value is None else str(cell.value)

            # If the cell has newlines, size to the longest line
            longest = max((len(line) for line in val.splitlines()), default=0)

            widths[col_letter] = max(widths.get(col_letter, 0), longest)

    for col, w in widths.items():
        ws.column_dimensions[col].width = max(min(w + padding, max_width), min_width)


def summarize_interfaces(data):
    result = {}
    for intf in data:
        attrs = intf.get("l1PhysIf", {}).get("attributes", {})
        dn = attrs.get("dn")
        state = attrs.get("operSt")
        if dn and state:
            result[dn] = state
    return result


def summarize_interface_errors(interface_errors):
    summary = {}
    for entry in interface_errors:
        dn = entry.get("dn")
        crc = int(entry.get("crc", 0))
        input_discards = int(entry.get("inputDiscards", 0))
        total_errors = crc + input_discards
        if dn:
            summary[dn] = total_errors
    return summary


def extract_interface_from_dn(dn):
    """
    Extract node ID and port from DN string.
    Example input: "topology/pod-1/node-102/sys/phys-[eth1/5]/dbgEtherStats"
    Output: ("node-102", "eth1/5")
    """
    match = re.search(r"node-(\d+).*phys-\[(.*?)\]", dn)
    if match:
        node_id = f"node-{match.group(1)}"
        port = match.group(2)
        return node_id, port
    return None, None


def write_interface_errors(intf_ws, category: str, changes: list):
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for entry in changes:
        node = entry["node"]
        interface = entry["interface"]
        intf_descr = entry["interface_descr"]
        before = entry["before"]
        after = entry["after"]
        endpoints = entry.get("endpoints", [])

        start_row = intf_ws.max_row + 1

        if endpoints:
            for ep in endpoints:
                intf_ws.append(
                    [
                        category,
                        node,
                        interface,
                        intf_descr,
                        before,
                        after,
                        ep.get("mac", ""),
                        ep.get("ip", ""),
                        ep.get("vlan", ""),
                        ep.get("epg_descr", ""),
                    ]
                )
            end_row = intf_ws.max_row

            for col in ("A", "B", "C", "D", "E", "F"):
                intf_ws.merge_cells(f"{col}{start_row}:{col}{end_row}")
                intf_ws[f"{col}{start_row}"].alignment = align_center

        else:
            intf_ws.append(
                [category, node, interface, intf_descr, before, after, "", "", "", ""]
            )
            r = intf_ws.max_row
            # Align the entire single row
            for col in ("A", "B", "C", "D", "E", "F", "G", "H"):
                intf_ws[f"{col}{r}"].alignment = align_center

    _autosize_columns(intf_ws)


def compare_snapshots(file1, file2):
    devices = load_devices()
    apics = []
    for device in devices:
        apics.append(device.get("hostname", ""))

    with open(file1) as f1, open(file2) as f2:
        before_json = json.load(f1)
        after_json = json.load(f2)

    result = {}

    for apic in apics:
        before = before_json.get(apic, {})
        after = after_json.get(apic, {})

        if not before or not after:
            continue

        result[apic] = {}

        result[apic]["fabric_health"] = {
            "before": before.get("fabric_health"),
            "after": after.get("fabric_health"),
        }

        # Faults
        before_faults = {
            f["faultInst"]["attributes"]["dn"] for f in before.get("faults", [])
        }
        after_faults = {
            f["faultInst"]["attributes"]["dn"] for f in after.get("faults", [])
        }

        result[apic]["new_faults"] = sorted(after_faults - before_faults)
        result[apic]["cleared_faults"] = sorted(before_faults - after_faults)

        # Endpoints
        before_eps = {
            ep["mac"]: {
                "node": ep["node"],
                "interface": ep["interface"],
                "mac": ep["mac"],
                "vlan": ep["vlan"],
                "epg_descr": ep["epg_descr"],
                "dn": ep["dn"],
                "ip": (ep.get("ip") or ""),
            }
            for ep in before.get("endpoints", [])
        }
        after_eps = {
            ep["mac"]: {
                "node": ep["node"],
                "interface": ep["interface"],
                "mac": ep["mac"],
                "vlan": ep["vlan"],
                "epg_descr": ep["epg_descr"],
                "dn": ep["dn"],
                "ip": (ep.get("ip") or ""),
            }
            for ep in after.get("endpoints", [])
        }

        before_dns = set(before_eps.keys())
        after_dns = set(after_eps.keys())
        new_endpoints = []
        missing_endpoints = []
        new_endpoints_mac = sorted(after_dns - before_dns)
        missing_endpoints_mac = sorted(before_dns - after_dns)
        for mac in new_endpoints_mac:
            temp = after_eps.get(mac, {})
            new_endpoints.append(temp)

        for mac in missing_endpoints_mac:
            temp = before_eps.get(mac)
            missing_endpoints.append(temp)

        result[apic]["new_endpoints"] = new_endpoints
        result[apic]["missing_endpoints"] = missing_endpoints

        # Interface status
        before_intfs = summarize_interfaces(before.get("interfaces", []))
        after_intfs = summarize_interfaces(after.get("interfaces", []))
        intf_changes = {
            "status_changed": [
                f"{k}: {before_intfs[k]} ➜ {after_intfs[k]}"
                for k in before_intfs.keys() & after_intfs.keys()
                if before_intfs[k] != after_intfs[k]
            ],
            "missing": sorted(set(before_intfs) - set(after_intfs)),
            "new": sorted(set(after_intfs) - set(before_intfs)),
        }
        result[apic]["interface_changes"] = intf_changes

        # NOTE: Interface Errors
        interfaces_map = {}
        po_map = {}
        interfaces = after.get("interfaces", [])
        pos = after.get("pc_aggr", [])

        for item in interfaces:
            attr = item.get("l1PhysIf", {}).get("attributes", {})

            dn = attr.get("dn", "")

            node = None
            for part in dn.split("/"):
                if part.startswith("node-"):
                    node = part
                    break

            entry = {
                "node": node,
                "id": attr.get("id", ""),
                "descr": attr.get("descr", ""),
            }
            interfaces_map[f"{node}-{attr.get('id', 'none')}"] = entry

            for po in pos:
                attr = po.get("pcAggrIf", {}).get("attributes", {})
                dn = attr.get("dn", "")
                node = None
                for part in dn.split("/"):
                    if part.startswith("node-"):
                        node = part
                        break

                po_entry = {
                    "node": node,
                    "id": attr.get("id", ""),
                    "name": attr.get("name", ""),
                }

                po_map[f"{node}-{attr.get('id', 'None')}"] = po_entry

            before_errs = summarize_interface_errors(before.get("interface_errors", []))
            after_errs = summarize_interface_errors(after.get("interface_errors", []))
            error_changes = {}
            for dn in set(before_errs) | set(after_errs):
                b = before_errs.get(dn, 0)
                a = after_errs.get(dn, 0)
                if a > b:
                    error_changes[dn] = f"{b} ➜ {a}"
            result[apic]["interface_error_changes"] = error_changes

        # NOTE: CRC Errors
        before_crc = {}
        for e in before.get("crc_errors", []):
            if "rmonEtherStats" in e and "attributes" in e["rmonEtherStats"]:
                dn = e["rmonEtherStats"]["attributes"].get("dn")
                # Note: The key is "cRCAlignErrors" not "crcAlignErrors"
                crc_align_errors = int(
                    e["rmonEtherStats"]["attributes"].get("cRCAlignErrors", 0)
                )
                if dn:
                    before_crc[dn] = crc_align_errors

        after_crc = {}
        for e in after.get("crc_errors", []):
            if "rmonEtherStats" in e and "attributes" in e["rmonEtherStats"]:
                dn = e["rmonEtherStats"]["attributes"].get("dn")
                # Note: The key is "cRCAlignErrors" not "crcAlignErrors"
                crc_align_errors = int(
                    e["rmonEtherStats"]["attributes"].get("cRCAlignErrors", 0)
                )
                if dn:
                    after_crc[dn] = crc_align_errors

        all_interfaces = set(before_crc.keys()) | set(after_crc.keys())
        crc_changes = []
        for dn in all_interfaces:
            b = before_crc.get(dn, 0)
            a = after_crc.get(dn, 0)

            if a > b:
                # Extract interface name for better readability
                node, port = extract_interface_from_dn(dn)
                if node is None:
                    continue

                err_eps = [
                    ep
                    for ep in after_eps.values()
                    if ep["node"] == node.split("-")[1] and ep["interface"] == port
                ]

                crc_changes.append(
                    {
                        "before": b,
                        "after": a,
                        "node": node,
                        "interface": port
                        if not port.startswith("po")  # type: ignore
                        else po_map.get(f"{node}-{port}", {}).get("name", port),
                        "interface_descr": interfaces_map.get(f"{node}-{port}", {}).get(
                            "descr", ""
                        ),
                        "endpoints": err_eps,
                    }
                )

        result[apic]["crc_error_changes"] = crc_changes

        # NOTE: Interface Errors
        before_drop = {}
        for e in before.get("drop_errors", []):
            if "rmonEgrCounters" in e and "attributes" in e["rmonEgrCounters"]:
                dn = e["rmonEgrCounters"]["attributes"].get("dn")
                drop_errors = int(
                    e["rmonEgrCounters"]["attributes"].get("bufferdroppkts", 0)
                )
                if dn:
                    before_drop[dn] = drop_errors
        after_drop = {}
        for e in after.get("drop_errors", []):
            if "rmonEgrCounters" in e and "attributes" in e["rmonEgrCounters"]:
                dn = e["rmonEgrCounters"]["attributes"].get("dn")
                drop_errors = int(
                    e["rmonEgrCounters"]["attributes"].get("bufferdroppkts", 0)
                )
                if dn:
                    after_drop[dn] = drop_errors

        all_interfaces = set(before_drop.keys()) | set(after_drop.keys())
        drop_changes = []
        for dn in all_interfaces:
            b = before_drop.get(dn, 0)
            a = after_drop.get(dn, 0)
            if a > b:
                node, intf = extract_interface_from_dn(dn)
                if node is None:
                    continue
                err_eps = [
                    ep
                    for ep in after_eps.values()
                    if ep["node"] == node.split("-")[1] and ep["interface"] == intf
                ]

                drop_changes.append(
                    {
                        "before": b,
                        "after": a,
                        "node": node,
                        "interface": intf
                        if not intf.startswith("po")  # type: ignore
                        else po_map.get(f"{node}-{intf}", {}).get("name", intf),
                        "interface_descr": interfaces_map.get(f"{node}-{intf}", {}).get(
                            "descr", ""
                        ),
                        "endpoints": err_eps,
                    }
                )

        result[apic]["drop_error_changes"] = drop_changes

        # NOTE: Output Errrors
        before_output = {}
        for e in before.get("output_errors", []):
            if "rmonIfOut" in e and "attributes" in e["rmonIfOut"]:
                dn = e["rmonIfOut"]["attributes"].get("dn")
                output_errors = int(e["rmonIfOut"]["attributes"].get("errors", 0))
                if dn:
                    before_output[dn] = output_errors
        after_output = {}
        for e in after.get("output_errors", []):
            if "rmonIfOut" in e and "attributes" in e["rmonIfOut"]:
                dn = e["rmonIfOut"]["attributes"].get("dn")
                output_errors = int(e["rmonIfOut"]["attributes"].get("errors", 0))
                if dn:
                    after_output[dn] = output_errors

        output_changes = []
        all_interfaces = set(before_output.keys()) | set(after_output.keys())
        for dn in all_interfaces:
            b = before_output.get(dn, 0)
            a = after_output.get(dn, 0)
            if a > b:
                node, intf = extract_interface_from_dn(dn)

                # === PATCH: Safe interface handling ===
                if intf is None:
                    safe_interface = ""
                else:
                    safe_interface = (
                        po_map.get(f"{node}-{intf}", {}).get("name", intf)
                        if intf.startswith("po")
                        else intf
                    )
                # =======================================

                err_eps = [
                    ep
                    for ep in after_eps.values()
                    if ep["node"] == node and ep["interface"] == intf
                ]

                output_changes.append(
                    {
                        "before": b,
                        "after": a,
                        "node": node,
                        "interface": safe_interface,
                        "interface_descr": interfaces_map.get(f"{node}-{intf}", {}).get(
                            "descr", ""
                        ),
                        "endpoints": err_eps,
                    }
                )

        result[apic]["output_error_changes"] = output_changes

        # NOTE: URIB Route
        before_routes = {
            r["uribv4Route"]["attributes"]["dn"] for r in before.get("urib_routes", [])
        }
        after_routes = {
            r["uribv4Route"]["attributes"]["dn"] for r in after.get("urib_routes", [])
        }
        route_changes = {
            "missing": sorted(before_routes - after_routes),
            "new": sorted(after_routes - before_routes),
        }
        result[apic]["urib_route_changes"] = route_changes

    return result


def print_colored_result(result):
    rprint("\n📈 [bold]COMPARISON RESULT:[/bold]\n")

    # Print summary counts
    rprint("[bold underline]Summary:[/bold underline]")
    for section, content in result.items():
        if section == "fabric_health":
            continue
        if isinstance(content, dict):
            count = len(content)
        elif isinstance(content, list):
            count = len(content)
        else:
            count = 1
        rprint(f"• [cyan]{section}[/cyan]: [bold yellow]{count}[/bold yellow]")
    rprint("")

    def print_section(title, content):
        rprint(f"🔹 [cyan]{title}[/cyan]:")
        if isinstance(content, dict):
            if not content:
                rprint("  (none)")
            else:
                for k, v in content.items():
                    rprint(f"  • {k}: {v}")
        elif isinstance(content, list):
            if not content:
                rprint("  (none)")
            else:
                for item in content:
                    rprint(f"  • {item}")
        else:
            rprint(f"  {content}")
        rprint("")  # spacing

    for section in [
        "fabric_health",
        "new_faults",
        "cleared_faults",
        "new_endpoints",
        "missing_endpoints",
        "moved_endpoints",
        "interface_changes",
        "interface_error_changes",
        "crc_error_changes",
        "drop_error_changes",
        "output_error_changes",
        "urib_route_changes",
    ]:
        if section in result:
            print_section(section, result[section])
        else:
            rprint(f"🔹 [yellow]{section}[/yellow]: (not available)\n")


def save_to_excel(all_result: dict, filename=None, base_dir=None):
    devices = load_devices()
    apics = [device.get("hostname", "") for device in devices]
    customer_name = get_customer_name()
    # Create directory structure
    if base_dir:
        compare_dir = os.path.join(base_dir, customer_name, "aci", "compare")
    else:
        compare_dir = os.path.join("results", customer_name, "aci", "compare")
    os.makedirs(compare_dir, exist_ok=True)

    if filename is None:
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        filename = f"{customer_name}_comparison_result_{timestamp}.xlsx"

    filepath = os.path.join(compare_dir, filename)

    wb = Workbook()
    wb.remove(wb.active)  # type: ignore

    for apic in apics:
        result = all_result.get(apic, {}) or {}
        ws = wb.create_sheet(f"{apic} - general")

        ws.append(["Category", "Item", "Details"])

        fh = result.get("fabric_health", {})
        ws.append(["Fabric Health", "Before", fh.get("before", "N/A")])
        ws.append(["Fabric Health", "After", fh.get("after", "N/A")])

        # Faults
        for fault in result.get("new_faults", []):
            ws.append(["New Faults", fault, ""])
        for fault in result.get("cleared_faults", []):
            ws.append(["Cleared Faults", fault, ""])

        # Interface Changes
        intf_changes = result.get("interface_changes", {}) or {}
        for change in intf_changes.get("status_changed", []):
            ws.append(["Interface Status Changed", change, ""])
        for intf in intf_changes.get("missing", []):
            ws.append(["Interface Missing", intf, ""])
        for intf in intf_changes.get("new", []):
            ws.append(["Interface New", intf, ""])

        # Interface Error Changes (dict or list supported)
        iec = result.get("interface_error_changes", {}) or {}
        if isinstance(iec, dict):
            for k, v in iec.items():
                ws.append(["Interface Error Changes", str(k), v])
        elif isinstance(iec, list):
            for row in iec:
                dn = (
                    row.get("dn") or f"{row.get('node', '')}/{row.get('interface', '')}"
                )
                before = row.get("before", "")
                after = row.get("after", "")
                ws.append(["Interface Error Changes", dn, f"{before} -> {after}"])

        # URIB Route Changes
        urib = result.get("urib_route_changes", {}) or {}
        for route in urib.get("missing", []):
            ws.append(["URIB Routes Missing", route, ""])
        for route in urib.get("new", []):
            ws.append(["URIB Routes New", route, ""])

        _autosize_columns(ws)

        # NOTE: Interface Error Sheet
        int_sheet_name = f"{apic} - Interface Errors"
        intf_ws = wb.create_sheet(int_sheet_name[:31])

        headers = [
            "Category",
            "Node",
            "Interface",
            "Description",
            "Before",
            "After",
            "Mac",
            "IP",
            "VLAN",
            "Description",
        ]
        intf_ws.append(headers)

        cec = result.get("crc_error_changes", [])
        dec = result.get("drop_error_changes", [])
        oec = result.get("output_error_changes", [])

        write_interface_errors(intf_ws, "CRC Changes", cec)
        write_interface_errors(intf_ws, "Drop Changes", dec)
        write_interface_errors(intf_ws, "Output Error Changes", oec)

        # NOTE: Endpoints Sheet
        eps_sheet_name = f"{apic} - Endpoints"
        eps_ws = wb.create_sheet(eps_sheet_name[:31])

        eps_ws.append(
            [
                "Category",
                "Endpoint",
                "Mac",
                "Ip address",
                "Switch",
                "interface",
                "vlan",
                "Description",
            ]
        )

        for cat, key in (
            ("New Endpoints", "new_endpoints"),
            ("Missing Endpoints", "missing_endpoints"),
        ):
            for ep in result.get(key, []) or []:
                if not isinstance(ep, dict):
                    continue
                eps_ws.append(
                    [
                        cat,
                        ep.get("dn", ""),
                        ep.get("mac", ""),
                        ep.get("ip", ""),
                        ep.get("node", ""),
                        ep.get("interface", ""),
                        ep.get("vlan", ""),
                        ep.get("epg_descr", ""),
                    ]
                )

            _autosize_columns(eps_ws)

        wb.save(filepath)

def compare_select(base_dir):
    print("\n📂 Selecting snapshots to compare...")
    file1, file2 = choose_snapshots(base_dir)
    if file1 and file2:
        print(f"📊 Comparing '{file1}' and '{file2}'...")
        result = compare_snapshots(file1, file2)
        # print_colored_result(result)
        save_to_excel(result, base_dir=base_dir)
        print("✅ Comparison results saved to Excel.")
    else:
        print("❌ No valid snapshots selected.")        

def compare_last_two(base_dir):
    customer_name = get_customer_name()
    if base_dir:
        print("base dir")
        files = sorted(
            glob.glob(f"{base_dir}/{customer_name}/aci/snapshot/*_snapshot_*.json")
        )
    else:
        files = sorted(
            glob.glob(f"results/{customer_name}/aci/snapshot/*_snapshot_*.json")
        )
    if len(files) < 2:
        print("❌ Not enough snapshot files found to compare.")
    else:
        before, after = files[-2], files[-1]
        print(f"📊 Comparing:\n  BEFORE: {before}\n  AFTER:  {after}")
        result = compare_snapshots(before, after)
        save_to_excel(result, base_dir=base_dir)
        print("✅ Comparison results saved to Excel.")
