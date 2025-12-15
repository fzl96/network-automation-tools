from netmiko import ConnectHandler
from netmiko import redispatch
import time
import re
import os
import getpass
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
try:
    from prettytable import PrettyTable
except ImportError:
    PrettyTable = None  # prettytable optional
import sys
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")))
from sp_tools.jumphost import get_jumpserver


#### Login to router ####

def interactive_main():
    banner = (f"""  
     ______________________________________________________________________
    |:..                                                   ''':::::%%%%%%%%|
    |%%%:::::..                      C R C E L L                 ''::::%%%%|
    |%%%%%%%:::::.....__________________________________________________:::|
      
    Description   : Automated CRC Error Monitoring with Delta Calculation
    Version       : 1.0
    ------------------------------------------------------------------------
    """)
    print(banner)
    
    jumpserver = get_jumpserver()
    #---------------------------------------------------------------------------------------------------------------------
    
    #---------------------------------------------------------------------------------------------------------------------
    # Here, the script connects to the jump server and confirms the connection by printing the prompt.
    # Connect to Jumphost
    print(f"\n{'~'*50}\n1. Connecting to the Jump host\n{'~'*50}")
    net_connect = ConnectHandler(**jumpserver)
    print ("Jump Server Prompt: {}".format(net_connect.find_prompt()))
    print(net_connect.find_prompt())
    print(f"\n{'*'*10}Connected to the Jump host{'*'*10}")
    #---------------------------------------------------------------------------------------------------------------------
    
    #---------------------------------------------------------------------------------------------------------------------
    # This section initiates an SSH connection to the destination router and reads the output.
    # Connect to Router
    print(f"\n{'~'*50}\n2. Connecting to the Dest Router\n{'~'*50}")
    #net_connect.write_channel("telnet p-d2-boo\n")
    destination = input("Enter the destination to connect (e.g: 'p-d2-bks', 'me-d6-sbr-4'): ")
    connection_type = "telnet"
    remote_node = f"{connection_type} {destination}"
    net_connect.write_channel(f"{remote_node}\n")
    time.sleep(3)
    output = net_connect.read_channel()
    print(output)
    #---------------------------------------------------------------------------------------------------------------------
    
    #---------------------------------------------------------------------------------------------------------------------
    # 3. kalau diminta username/password, kirim kredensial langsung dari user input
    if "username" in output.lower():
        router_username = input("Enter router username: ").strip()
        router_password = getpass.getpass("Enter router password: ")

        net_connect.write_channel(f"{router_username}\n")
        time.sleep(2)
        net_connect.write_channel(f"{router_password}\n")

        print(f"\n{'~'*50}\n3. Destination Device Prompt\n{'~'*50}\n")
        print(net_connect.find_prompt())
        print("Router Prompt: {}".format(net_connect.find_prompt()))

        redispatch(net_connect, device_type="cisco_ios")
    
    # Print timestamp with day
    #print(time.strftime("%a %d-%b-%Y %H:%M:%S WIB", time.localtime()))
    
    
    
    
    ####---------------------------------------------------------------------------------------------------------------------#####
    
    
    
    # This section sends a command to the router and prints the output.
    print(f"\n{'~'*50}\n5. Sending commands to the router\n{'~'*50}")
    
    termlen_output = net_connect.send_command("terminal length 0")
    print("Output for command 'term len 0':\n", termlen_output)
    
    # Send command to show interfaces and filter out admin-down interfaces
    show_int = net_connect.send_command("show interfaces hundredGigE * brief | exclude admin-down")
    print("Output for command 'show interfaces hundredGigE * brief | exclude admin-down':\n", show_int)
    
    # parse interface using regex
    def parse_interface(data):
        return re.findall(r'\b(?:Hu)\S*\b', data)
    
    # create command "show interface {interface} include CRC"
    def generate_commands(parsed_interfaces):
        commands = []
        for interface in parsed_interfaces:
            command = f"show interface {interface} | include CRC"
            commands.append(command)
        return commands
    
    # Generate commands for each interface
    interfaces = parse_interface(show_int)
    if not interfaces:
        print("WARNING: No HundredGigE interfaces were found. The script will terminate.")
        net_connect.disconnect()
        return
    
    commands = generate_commands(interfaces)
    
    # A dictionary to store all attempts, keyed by interface name
    interface_attempts = {interface: [] for interface in interfaces}
    
    # New Loop structure: Perform all attempts for each interface sequentially
    # This section has been modified to handle the new looping logic
    print(f"\n{'~'*50}\n5a. Executing commands for all interfaces (Attempt 1)\n{'~'*50}")
    for command in commands:
        interface_name = re.search(r'\b(?:Hu)\S*\b', command).group(0)
        print(f"Executing command: '{command}' (Attempt 1)")
        show_crc = net_connect.send_command(command)
        
        # Parse the output and store
        input_match = re.search(r'(\d+)\s+input\s+errors', show_crc, re.IGNORECASE)
        crc_match = re.search(r'(\d+)\s+CRC', show_crc, re.IGNORECASE)
        output_match = re.search(r'(\d+)\s+output\s+errors', show_crc, re.IGNORECASE)
    
        input_errors = int(input_match.group(1)) if input_match else 0
        crc = int(crc_match.group(1)) if crc_match else 0
        output_errors = int(output_match.group(1)) if output_match else 0
        
        interface_attempts[interface_name].append({
            "input_errors": input_errors,
            "CRC": crc,
            "output_errors": output_errors
        })
        print(f"Command output (Attempt 1):\n{show_crc}\n")
    
    # Delay before attempt 2
    print("Waiting 2 second before starting Attempt 2...")
    time.sleep(2)
    
    print(f"\n{'~'*50}\n5b. Executing commands for all interfaces (Attempt 2)\n{'~'*50}")
    for command in commands:
        interface_name = re.search(r'\b(?:Hu)\S*\b', command).group(0)
        print(f"Executing command: '{command}' (Attempt 2)")
        show_crc = net_connect.send_command(command)
        
        # Parse the output and store
        input_match = re.search(r'(\d+)\s+input\s+errors', show_crc, re.IGNORECASE)
        crc_match = re.search(r'(\d+)\s+CRC', show_crc, re.IGNORECASE)
        output_match = re.search(r'(\d+)\s+output\s+errors', show_crc, re.IGNORECASE)
    
        input_errors = int(input_match.group(1)) if input_match else 0
        crc = int(crc_match.group(1)) if crc_match else 0
        output_errors = int(output_match.group(1)) if output_match else 0
        
        interface_attempts[interface_name].append({
            "input_errors": input_errors,
            "CRC": crc,
            "output_errors": output_errors
        })
        print(f"Command output (Attempt 2):\n{show_crc}\n")
    
    # Delay before attempt 3
    print("Waiting 2 seconds before starting Attempt 3...")
    time.sleep(2)
    
    print(f"\n{'~'*50}\n5c. Executing commands for all interfaces (Attempt 3)\n{'~'*50}")
    for command in commands:
        interface_name = re.search(r'\b(?:Hu)\S*\b', command).group(0)
        print(f"Executing command: '{command}' (Attempt 3)")
        show_crc = net_connect.send_command(command)
        
        # Parse the output and store
        input_match = re.search(r'(\d+)\s+input\s+errors', show_crc, re.IGNORECASE)
        crc_match = re.search(r'(\d+)\s+CRC', show_crc, re.IGNORECASE)
        output_match = re.search(r'(\d+)\s+output\s+errors', show_crc, re.IGNORECASE)
    
        input_errors = int(input_match.group(1)) if input_match else 0
        crc = int(crc_match.group(1)) if crc_match else 0
        output_errors = int(output_match.group(1)) if output_match else 0
        
        interface_attempts[interface_name].append({
            "input_errors": input_errors,
            "CRC": crc,
            "output_errors": output_errors
        })
        print(f"Command output (Attempt 3):\n{show_crc}\n")
    
    
    # --- Data Extraction and Table Generation (Delta Calculation) ---
    print(f"\n{'~'*50}\n6. Processing Data and Generating Table\n{'~'*50}")
    
    # Prepare a list to hold the delta data
    delta_data = []
    
    for interface, attempts in interface_attempts.items():
        if len(attempts) >= 3:
            # Get data from attempt 1 and attempt 3
            attempt1 = attempts[0]
            attempt3 = attempts[2]
            
            # Calculate the delta (selisih) for each metric
            delta_input_errors = attempt3["input_errors"] - attempt1["input_errors"]
            delta_crc = attempt3["CRC"] - attempt1["CRC"]
            delta_output_errors = attempt3["output_errors"] - attempt1["output_errors"]
            
            delta_data.append({
                "interface": interface,
                "timestamp": datetime.now().strftime("%a %d-%b-%Y %H:%M:%S WIB"),
                "delta_input_errors": delta_input_errors,
                "delta_crc": delta_crc,
                "delta_output_errors": delta_output_errors
            })
        else:
            print(f"WARNING: Insufficient attempts for interface {interface}. Skipped delta calculation.")
    
    # Create table string for output (PrettyTable jika tersedia)
    if PrettyTable is not None:
        # gunakan prettytable jika modul tersedia
        table = PrettyTable()
        table.field_names = ["Interface", "Timestamp", "Delta Input Errors", "Delta CRC", "Delta Output Errors"]
        
        for entry in delta_data:
            table.add_row([
                entry["interface"],
                entry["timestamp"],
                entry["delta_input_errors"],
                entry["delta_crc"],
                entry["delta_output_errors"]
            ])
        
        table_str = str(table)
    else:
        # fallback jika prettytable tidak terinstall
        header = "Interface, Timestamp, Delta Input Errors, Delta CRC, Delta Output Errors"
        rows = []
        for entry in delta_data:
            rows.append(
                f"{entry['interface']}, "
                f"{entry['timestamp']}, "
                f"{entry['delta_input_errors']}, "
                f"{entry['delta_crc']}, "
                f"{entry['delta_output_errors']}"
            )
        table_str = header + "\n" + "\n".join(rows)
        print("WARNING: prettytable not installed, using simple text table.\n")

    print(table_str)
    
    # --- File Saving .txt (Dipertahankan untuk cadangan) ---
    print(f"\n{'~'*50}\n7. Writing output to file\n{'~'*50}")
    timestamp = datetime.now().strftime("%d-%b-%Y_%H-%M-%S_WIB")
    txt_filename = f'CRC_DELTA_{timestamp}.txt'
    with open(txt_filename, 'w') as f:
        f.write(table_str)
    print(f"Data saved to text file: {txt_filename}")
    
    
    # --- File Saving .xlsx (Menyimpan, meng-append, dan mewarnai) ---
    excel_filename = 'CRC_DELTA_SUMMARY.xlsx'
    sheet_name = destination
    red_font = Font(color="FF0000") # Definisi font merah
    
    try:
        df_new = pd.DataFrame(delta_data)
        sheets_data = {}
        
        if os.path.exists(excel_filename):
            sheets_data = pd.read_excel(excel_filename, sheet_name=None)
            if sheet_name in sheets_data:
                df_old = sheets_data[sheet_name]
                df_combined = pd.concat([df_old, df_new], ignore_index=True)
                sheets_data[sheet_name] = df_combined
            else:
                sheets_data[sheet_name] = df_new
        else:
            sheets_data[sheet_name] = df_new
    
        # Menulis semua data ke Excel
        with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
            for sheet_title, df_content in sheets_data.items():
                df_content.to_excel(writer, sheet_name=sheet_title, index=False)
        
        # --- Tambahan untuk Pewarnaan Font ---
        print(f"Applying conditional formatting to sheet '{sheet_name}'...")
        wb = load_workbook(excel_filename)
        ws = wb[sheet_name]
    
        # Cari nomor kolom untuk setiap header
        col_map = {header: idx + 1 for idx, header in enumerate(ws[1])}
        
        col_input = col_map.get("delta_input_errors")
        col_crc = col_map.get("delta_crc")
        col_output = col_map.get("delta_output_errors")
    
        # Iterasi dari baris kedua (data, bukan header) hingga akhir
        for row in range(2, ws.max_row + 1):
            if col_input and ws.cell(row=row, column=col_input).value > 0:
                ws.cell(row=row, column=col_input).font = red_font
            if col_crc and ws.cell(row=row, column=col_crc).value > 0:
                ws.cell(row=row, column=col_crc).font = red_font
            if col_output and ws.cell(row=row, column=col_output).value > 0:
                ws.cell(row=row, column=col_output).font = red_font
        
        # Menyimpan file Excel dengan format yang sudah diterapkan
        wb.save(excel_filename)
        
        print(f"Data saved and formatted in Excel file: {excel_filename} on sheet '{sheet_name}'")
    
    except ImportError:
        print("Warning: pandas or openpyxl not installed. Could not save to Excel.")
    
    
    # Disconnect from the device
    print(f"\n{'~'*50}\n8. Disconnecting from device\n{'~'*50}")
    net_connect.disconnect()


def run_crc_gui(jump_ip, username, password, port=22, destination=None):
    # Wrapper for GUI usage: patch get_jumpserver() and input(), then run interactive_main()
    import builtins
    import io
    import contextlib

    # Normalisasi nilai input
    jump_ip = (jump_ip or "").strip()
    username = (username or "").strip()
    password = password or ""
    destination = (destination or "").strip()
    port = int(port) if port else 22

    if not jump_ip or not username or not password:
        raise ValueError("Jumpserver IP / username / password tidak boleh kosong")
    if not destination:
        raise ValueError("Destination router tidak boleh kosong")

    global get_jumpserver
    old_get = get_jumpserver

    def fake_get_jumpserver():
        # Meniru format return get_jumpserver() asli
        return {
            "device_type": "terminal_server",
            "ip": jump_ip,
            "username": username,
            "password": password,
            "port": port,
        }

    get_jumpserver = fake_get_jumpserver

    answers = []
    if destination:
        answers.append(destination)

    old_input = builtins.input

    def fake_input(prompt=""):
        if answers:
            return answers.pop(0)
        return ""

    builtins.input = fake_input

    buf = io.StringIO()
    try:
        with contextlib.redirect_stdout(buf):
            interactive_main()
    finally:
        get_jumpserver = old_get
        builtins.input = old_input

    return buf.getvalue()


if __name__ == "__main__":
    interactive_main()
