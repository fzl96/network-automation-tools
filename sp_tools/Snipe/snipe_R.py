from netmiko import ConnectHandler
from netmiko import redispatch
import time
import re
try:
    from prettytable import PrettyTable
except ImportError:
    PrettyTable = None  # prettytable optional
import datetime
import os
from openpyxl import Workbook, load_workbook
import sys
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")))
from sp_tools.jumphost import get_jumpserver

#### Login to router ####

def interactive_main():
    banner = (f"""
        
     _________________________________________________________________________________________
    |:..                                                                     ''':::::%%%%%%%%|
    |%%%:::::..                               S N I P E                            ''::::%%%%|
    |%%%%%%%:::::.....____________________________________________________________________:::|
      
    Description   : Automated IOS-XR protocols count in summary to Excel
    Version       : 1.0
    -----------------------------------------------------------------------------------------
    
    """)
    print(banner)
    
    jumpserver = get_jumpserver()
    #---------------------------------------------------------------------------------------------------------------------
    
    #---------------------------------------------------------------------------------------------------------------------
    # Here, the script connects to the jump server and confirms the connection by printing the prompt.
    # Connect to Jumphost
    print(f"\n{'~'*50}\n1. Connecting to the Jump host\n{'~'*50}")
    net_connect = ConnectHandler(**jumpserver)
    print ("Jump Server Prompt: {}".format(net_connect.find_prompt()))
    print(net_connect.find_prompt())
    print(f"\n{'*'*10}Connected to the Jump host{'*'*10}")
    #---------------------------------------------------------------------------------------------------------------------
    
    #---------------------------------------------------------------------------------------------------------------------
    # This section initiates an SSH connection to the destination router and reads the output.
    # Connect to Router
    print(f"\n{'~'*50}\n2. Connecting to the Dest Router\n{'~'*50}")
    #net_connect.write_channel("telnet p-d2-boo\n")
    destination = input("Enter the destination to connect (e.g: 'p-d2-bks', 'me-d6-sbr-4'): ")
    connection_type = "telnet"
    remote_node = f"{connection_type} {destination}"
    net_connect.write_channel(f"{remote_node}\n")
    time.sleep(3)
    output = net_connect.read_channel()
    print(output)
    #---------------------------------------------------------------------------------------------------------------------
    
    #---------------------------------------------------------------------------------------------------------------------
    # If prompted for a password, the script sends the password and redispatches the connection to the appropriate device type.
    if "username" in output:
        # Read username and password from a .txt file
        with open('tacacs_credentials.txt', 'r') as cred_file:
            lines = cred_file.readlines()
            router_username = lines[0].strip()
            router_password = lines[1].strip()
        
        # Send username and password to the router
        net_connect.write_channel(f"{router_username}\n")
        time.sleep(2)
        net_connect.write_channel(f"{router_password}\n")
        print(f"\n{'~'*50}\n3. Destination Device Prompt\n{'~'*50}\n")
        print(net_connect.find_prompt())
        print("Router Prompt: {}".format(net_connect.find_prompt()))
        redispatch(net_connect, device_type="cisco_ios")
    
    # Print timestamp with day
    #print(time.strftime("%a %d-%b-%Y %H:%M:%S WIB", time.localtime()))
    
    
    ####---------------------------------------------------------------------------------------------------------------------
    # Get the current timestamp
    current_timestamp = datetime.datetime.now().strftime("%d/%m/%Y/%H:%M") # Format as DD/MM/YYYY/HH:MM
    
    # Send commands to the router and store the output
    ospf_raw = net_connect.send_command("show ospf neighbor | in FULL | utility wc line")
    ldp_raw = net_connect.send_command("show mpls ldp neighbor brief | in Y | utility wc line")
    cdp_raw = net_connect.send_command("show cdp neighbors | utility wc line")
    mpls_te_raw = net_connect.send_command("show mpls traffic-eng tunnels tabular | utility wc line")
    interface_up_raw = net_connect.send_command("show interfaces description | in up | utility wc line")
    interface_down_raw = net_connect.send_command("show interfaces description | in down | utility wc line")
    pim_raw = net_connect.send_command("show pim neighbor | utility wc line")
    
    # --- Applying your proven cleaning method ---
    # Remove lines containing "WIB or GMT" using regex and strip whitespace
    ospf_clean = re.sub(r'.*(WIB|GMT).*', '', ospf_raw, flags=re.MULTILINE).strip()
    ldp_clean = re.sub(r'.*(WIB|GMT).*', '', ldp_raw, flags=re.MULTILINE).strip()
    cdp_clean = re.sub(r'.*(WIB|GMT).*', '', cdp_raw, flags=re.MULTILINE).strip()
    mpls_te_clean = re.sub(r'.*(WIB|GMT).*', '', mpls_te_raw, flags=re.MULTILINE).strip()
    interface_up_clean = re.sub(r'.*(WIB|GMT).*', '', interface_up_raw, flags=re.MULTILINE).strip()
    interface_down_clean = re.sub(r'.*(WIB|GMT).*', '', interface_down_raw, flags=re.MULTILINE).strip()
    pim_clean = re.sub(r'.*(WIB|GMT).*', '', pim_raw, flags=re.MULTILINE).strip()
    
    # Safely convert to integer
    try:
        ospf_count = int(ospf_clean) if ospf_clean else 0
    except ValueError:
        ospf_count = 0 # Fallback if conversion fails
        print(f"Warning: Could not convert '{ospf_clean}' to int for OSPF count.")
    
    try:
        ldp_count = int(ldp_clean) if ldp_clean else 0
    except ValueError:
        ldp_count = 0 # Fallback if conversion fails
        print(f"Warning: Could not convert '{ldp_clean}' to int for LDP count.")
    
    try:
        cdp_count = int(cdp_clean) if cdp_clean else 0
    except ValueError:
        cdp_count = 0 # Fallback if conversion fails
        print(f"Warning: Could not convert '{cdp_clean}' to int for CDP count.")
    
    try:
        mpls_te_count = int(mpls_te_clean) if mpls_te_clean else 0
    except ValueError:
        mpls_te_count = 0 # Fallback if conversion fails
        print(f"Warning: Could not convert '{mpls_te_clean}' to int for MPLS-TE count.")
    
    try:
        interface_up_count = int(interface_up_clean) if interface_up_clean else 0
    except ValueError:
        interface_up_count = 0 # Fallback if conversion fails
        print(f"Warning: Could not convert '{interface_up_clean}' to int for Interface Up count.")
    
    try:
        interface_down_count = int(interface_down_clean) if interface_down_clean else 0
    except ValueError:  
        interface_down_count = 0 # Fallback if conversion fails
        print(f"Warning: Could not convert '{interface_down_clean}' to int for Interface Down count.")
    
    try:
        pim_count = int(pim_clean) if pim_clean else 0
    except ValueError:
        pim_count = 0 # Fallback if conversion fails
        print(f"Warning: Could not convert '{pim_clean}' to int for PIM count.")
    
    # --- End of cleaning method ---
    
    # Get the hostname (without the '#' or '>' prompt character)
    hostname = net_connect.find_prompt().strip('#').strip('>')
    # For Cisco XR (RP/0/RP0/CPU0:P-D2-BOO), extract just 'P-D2-BOO'
    if "RP/" in hostname and "CPU0:" in hostname:
        hostname = hostname.split(":")[-1].strip()  # Extract after the last colon
    
    # Create a table string for console output
    if PrettyTable is not None:
        # gunakan prettytable jika tersedia
        table = PrettyTable()
        table.field_names = ["Hostname", "Timestamp", "OSPF", "LDP", "CDP", "MPLS-TE", "Intf Up", "Intf Down", "PIM"]
        table.add_row([hostname, current_timestamp, ospf_count, ldp_count, cdp_count, mpls_te_count, interface_up_count, interface_down_count, pim_count])
        table_str = str(table)
    else:
        # fallback jika prettytable tidak ada
        header = "Hostname, Timestamp, OSPF, LDP, CDP, MPLS-TE, Intf Up, Intf Down, PIM"
        row = f"{hostname}, {current_timestamp}, {ospf_count}, {ldp_count}, {cdp_count}, {mpls_te_count}, {interface_up_count}, {interface_down_count}, {pim_count}"
        table_str = header + "\n" + row
        print("WARNING: prettytable not installed, using simple text table.\n")

    print(table_str)  # Print to console
    
    # Define the Excel file name
    excel_file = "snipe_table.xlsx"
    
    # Prepare the data row to be written to Excel
    data_row = [hostname, current_timestamp, ospf_count, ldp_count, cdp_count, mpls_te_count, interface_up_count, interface_down_count, pim_count]
    
    # Write and save the data to Excel file
    print(f"\n{'~'*50}\n4. Saving output to {excel_file}\n{'~'*50}\n")
    
    if not os.path.exists(excel_file):
        # If the file doesn't exist, create a new workbook
        wb = Workbook()
        # Remove the default sheet created by openpyxl
        ws = wb.active
        wb.remove(ws)
        
        # Create a new sheet with the name of the destination
        ws = wb.create_sheet(title=destination)
        headers = ["Hostname", "Timestamp", "OSPF", "LDP", "CDP", "MPLS-TE", "Intf Up", "Intf Down", "PIM"]
        ws.append(headers)
        print(f"Created new Excel file: {excel_file} with a new sheet '{destination}' and headers.")
    else:
        # If the file exists, load it
        wb = load_workbook(excel_file)
        print(f"Opened existing Excel file: {excel_file}.")
        
        # Check if a sheet with the destination's name already exists
        if destination in wb.sheetnames:
            ws = wb[destination]
            print(f"Opened existing sheet: '{destination}'.")
        else:
            # If the sheet doesn't exist, create a new one with headers
            ws = wb.create_sheet(title=destination)
            headers = ["Hostname", "Timestamp", "OSPF", "LDP", "CDP", "MPLS-TE", "Intf Up", "Intf Down", "PIM"]
            ws.append(headers)
            print(f"Created a new sheet '{destination}' with headers.")
    
    # Append the new data row to the correct sheet
    ws.append(data_row)
    print(f"Appended data for {hostname} at {current_timestamp}.")
    
    # Save the workbook
    wb.save(excel_file)
    print(f'Data has been saved to {excel_file}.\n')
    
    # Disconnect
    net_connect.disconnect()


def run_snipe_gui(jump_ip, username, password, port=22, destination=None):
    # Wrapper for GUI usage: patch get_jumpserver() and input(), then run interactive_main()
    import builtins
    import io
    import contextlib

    # Normalisasi nilai input
    jump_ip = (jump_ip or "").strip()
    username = (username or "").strip()
    password = password or ""
    destination = (destination or "").strip()
    port = int(port) if port else 22

    if not jump_ip or not username or not password:
        raise ValueError("Jumpserver IP / username / password tidak boleh kosong")
    if not destination:
        raise ValueError("Destination router tidak boleh kosong")

    global get_jumpserver
    old_get = get_jumpserver

    def fake_get_jumpserver():
        # Meniru format return get_jumpserver() asli
        return {
            "device_type": "terminal_server",
            "ip": jump_ip,
            "username": username,
            "password": password,
            "port": port,
        }

    get_jumpserver = fake_get_jumpserver

    answers = []
    if destination:
        answers.append(destination)

    old_input = builtins.input

    def fake_input(prompt=""):
        if answers:
            return answers.pop(0)
        return ""

    builtins.input = fake_input

    buf = io.StringIO()
    try:
        with contextlib.redirect_stdout(buf):
            interactive_main()
    finally:
        get_jumpserver = old_get
        builtins.input = old_input

    return buf.getvalue()


if __name__ == "__main__":
    interactive_main()
